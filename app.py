import os
import io
import sys
import subprocess
import hashlib
import zipfile
import pandas as pd
import streamlit as st
from datetime import date, datetime
from urllib.parse import quote as urlquote
from fpdf import FPDF

# ── Auto-install missing packages ────────────
def _ensure(pkg, import_name=None):
    import_name = import_name or pkg
    try:
        __import__(import_name)
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "-q"])

_ensure("openpyxl")
_ensure("fpdf2", "fpdf")
_ensure("supabase")

# ──────────────────────────────────────────────
# PAGE CONFIG
# ──────────────────────────────────────────────
st.set_page_config(
    page_title="ร้านบุญสุข Smart Sales v5",
    page_icon="❄️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ──────────────────────────────────────────────
# CONSTANTS
# ──────────────────────────────────────────────
STORE_NAME    = "ร้านบุญสุขอิเล็กทรอนิกส์"
STORE_PHONE   = "086-2613829"
STORE_WEB     = "https://www.facebook.com/boonsukele/"
STORE_ADDRESS = "87 หมู่ 12 ต.คาละแมะ อ.ศรีขรภูมิ จ.สุรินทร์ 32110"
STORE_TAX_ID  = ""   # ← ใส่เลขผู้เสียภาษีถ้ามี

INSTALL_CONDITIONS = (
    "1) แถมรางครอบท่อน้ำยาให้ฟรี ไม่เกิน 4 เมตร หากเกินคิดเพิ่ม เมตรละ 200 บาท\n"
    "2) แถมท่อน้ำยา ไม่เกิน 4 เมตร หากเกินคิดเพิ่ม เมตรละ 400 บาท\n"
    "3) แถมท่อน้ำทิ้ง ไม่เกิน 10 เมตร หากเกินคิดเพิ่ม เมตรละ 40 บาท\n"
    "4) แถมสายไฟ ไม่เกิน 10 เมตร หากเกินคิดเพิ่ม เมตรละ 40 บาท\n"
    "5) แถมขาแขวนหรือขายาง สำหรับติดตั้งคอยล์ร้อน\n"
    "6) กรณีไม่มีเบรคเกอร์ แถมให้ฟรี\n"
    "7) รับประกันงานตามเงื่อนไขฟรี ตลอดอายุการใช้งาน"
)

DATA_DIR  = "."
STOCK_CSV = os.path.join(DATA_DIR, "boonsuk_stock.csv")
LOG_CSV      = os.path.join(DATA_DIR, "boonsuk_customer_log.csv")
SERVICE_CSV  = os.path.join(DATA_DIR, "boonsuk_service_log.csv")

# ── ผู้ใช้งาน ──────────────────────────────────
# เปลี่ยนรหัสผ่านได้โดยแก้ค่าใน USERS
# สร้าง hash ใหม่: hashlib.sha256("รหัสผ่าน".encode()).hexdigest()
USERS = {
    "admin": hashlib.sha256("boonsuk2024".encode()).hexdigest(),
    "staff": hashlib.sha256("staff1234".encode()).hexdigest(),
}

JOB_STATUSES       = ["📋 รอดำเนินการ", "🔧 กำลังติดตั้ง", "✅ ติดตั้งแล้ว", "💰 รับเงินแล้ว", "❌ ยกเลิก"]
LOW_STOCK_THRESHOLD = 2   # แจ้งเตือนเมื่อสต๊อก ≤ ค่านี้

# ──────────────────────────────────────────────
# PRODUCT CATALOGUE
# ──────────────────────────────────────────────
PRODUCTS = [
    {"section":"Midea ฟิกส์speed","model":"Asmg09c","btu":9000,"price_install":19000,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"10 ปี","stock_qty":0},
    {"section":"Midea ฟิกส์speed","model":"Asmg12j","btu":12000,"price_install":21500,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"10 ปี","stock_qty":0},
    {"section":"Midea ฟิกส์speed","model":"Asaa18j","btu":18000,"price_install":27500,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"10 ปี","stock_qty":0},
    {"section":"Midea ฟิกส์speed","model":"Asaa24j","btu":24000,"price_install":37500,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"10 ปี","stock_qty":0},
    {"section":"Midea ฟิกส์speed","model":"Asaa30j","btu":30000,"price_install":43000,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"10 ปี","stock_qty":0},
    {"section":"Fujitsu DC Inverter iPower II R410A","model":"Asmg09jl","btu":8500,"price_install":15500,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"10 ปี","stock_qty":0},
    {"section":"Fujitsu DC Inverter iPower II R410A","model":"Asmg12jl","btu":11900,"price_install":16500,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"10 ปี","stock_qty":0},
    {"section":"Fujitsu DC Inverter iPower II R410A","model":"Asaa18jc","btu":17700,"price_install":23500,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"10 ปี","stock_qty":0},
    {"section":"Fujitsu DC Inverter iPower II R410A","model":"Asaa24jc","btu":24200,"price_install":34500,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"10 ปี","stock_qty":0},
    {"section":"Fujitsu DC Inverter iPower II R410A","model":"Asaa30cm","btu":27300,"price_install":40000,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"10 ปี","stock_qty":0},
    {"section":"Fujitsu Excellence Fix Speed R32","model":"Asma09r32","btu":9100,"price_install":13800,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"10 ปี","stock_qty":0},
    {"section":"Fujitsu Excellence Fix Speed R32","model":"Asma12r32","btu":11500,"price_install":14500,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"10 ปี","stock_qty":0},
    {"section":"Fujitsu Excellence Fix Speed R32","model":"Asma13r3","btu":13906,"price_install":16700,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"10 ปี","stock_qty":0},
    {"section":"Fujitsu Excellence Fix Speed R32","model":"Asma18r410","btu":18745,"price_install":23000,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"10 ปี","stock_qty":0},
    {"section":"Fujitsu Excellence Fix Speed R32","model":"Asma24r410","btu":24508,"price_install":32000,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"10 ปี","stock_qty":0},
    {"section":"Fujitsu Excellence Fix Speed R32","model":"Asma30r4","btu":28800,"price_install":35500,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"10 ปี","stock_qty":0},
    {"section":"Carrier Explorer Inverter","model":"Tvgs010","btu":9000,"price_install":15800,"w_install":"1 ปี","w_parts":"5 ปี","w_comp":"10 ปี","stock_qty":0},
    {"section":"Carrier Explorer Inverter","model":"Tvgs013","btu":12000,"price_install":18500,"w_install":"1 ปี","w_parts":"5 ปี","w_comp":"10 ปี","stock_qty":0},
    {"section":"Carrier Explorer Inverter","model":"Tvgs016","btu":15000,"price_install":22500,"w_install":"1 ปี","w_parts":"5 ปี","w_comp":"10 ปี","stock_qty":0},
    {"section":"Carrier Explorer Inverter","model":"Tvgs018","btu":18000,"price_install":26500,"w_install":"1 ปี","w_parts":"5 ปี","w_comp":"10 ปี","stock_qty":0},
    {"section":"Carrier Explorer Inverter","model":"Tvgs024","btu":22000,"price_install":29500,"w_install":"1 ปี","w_parts":"5 ปี","w_comp":"10 ปี","stock_qty":0},
    {"section":"Carrier Gemini Inverter","model":"Tvegb010","btu":9000,"price_install":15000,"w_install":"1 ปี","w_parts":"5 ปี","w_comp":"10 ปี","stock_qty":0},
    {"section":"Carrier Gemini Inverter","model":"Tvegb013","btu":12000,"price_install":17000,"w_install":"1 ปี","w_parts":"5 ปี","w_comp":"10 ปี","stock_qty":0},
    {"section":"Carrier Gemini Inverter","model":"Tvegb018","btu":18000,"price_install":23500,"w_install":"1 ปี","w_parts":"5 ปี","w_comp":"10 ปี","stock_qty":0},
    {"section":"Carrier Gemini Inverter","model":"Tvegb024","btu":22000,"price_install":27000,"w_install":"1 ปี","w_parts":"5 ปี","w_comp":"10 ปี","stock_qty":0},
    {"section":"Carrier Gemini Inverter","model":"Tvegb025","btu":24000,"price_install":31500,"w_install":"1 ปี","w_parts":"5 ปี","w_comp":"10 ปี","stock_qty":0},
    {"section":"Carrier Astrony R32","model":"AAF010","btu":9000,"price_install":13000,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"7 ปี","stock_qty":0},
    {"section":"Carrier Astrony R32","model":"AAF013","btu":12000,"price_install":14000,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"7 ปี","stock_qty":0},
    {"section":"Carrier Astrony R32","model":"AAF018","btu":18000,"price_install":20500,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"7 ปี","stock_qty":0},
    {"section":"Carrier Astrony R32","model":"AAF025","btu":25000,"price_install":26500,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"7 ปี","stock_qty":0},
    {"section":"Carrier Everest R32","model":"Tsgs010","btu":9000,"price_install":14000,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"7 ปี","stock_qty":0},
    {"section":"Carrier Everest R32","model":"Tsgs013","btu":12000,"price_install":15000,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"7 ปี","stock_qty":0},
    {"section":"Carrier Everest R32","model":"Tsgs018","btu":18000,"price_install":21000,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"7 ปี","stock_qty":0},
    {"section":"Carrier Everest R32","model":"Tsgs025","btu":24000,"price_install":27000,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"7 ปี","stock_qty":0},
    {"section":"Mitsubishi Heavy Duty Deluxe R32","model":"Srk10cvs","btu":9444,"price_install":15500,"w_install":"1 ปี","w_parts":"5 ปี","w_comp":"5 ปี","stock_qty":0},
    {"section":"Mitsubishi Heavy Duty Deluxe R32","model":"Srk13cvs","btu":12039,"price_install":18000,"w_install":"1 ปี","w_parts":"5 ปี","w_comp":"5 ปี","stock_qty":0},
    {"section":"Mitsubishi Heavy Duty Deluxe R32","model":"Srk19cvs","btu":19127,"price_install":29000,"w_install":"1 ปี","w_parts":"5 ปี","w_comp":"5 ปี","stock_qty":0},
    {"section":"Mitsubishi Heavy Duty Deluxe R32","model":"Srk25cvs","btu":25085,"price_install":38800,"w_install":"1 ปี","w_parts":"5 ปี","w_comp":"5 ปี","stock_qty":0},
    {"section":"Mitsubishi Heavy Duty Standard R32","model":"Srk10cvv","btu":9239,"price_install":15000,"w_install":"1 ปี","w_parts":"5 ปี","w_comp":"5 ปี","stock_qty":0},
    {"section":"Mitsubishi Heavy Duty Standard R32","model":"Srk13cvv","btu":11634,"price_install":17500,"w_install":"1 ปี","w_parts":"5 ปี","w_comp":"5 ปี","stock_qty":0},
    {"section":"Mitsubishi Heavy Duty Standard R32","model":"Srk15cvv","btu":14457,"price_install":20800,"w_install":"1 ปี","w_parts":"5 ปี","w_comp":"5 ปี","stock_qty":0},
    {"section":"Mitsubishi Heavy Duty Standard R32","model":"Srk18cvv","btu":17305,"price_install":25500,"w_install":"1 ปี","w_parts":"5 ปี","w_comp":"5 ปี","stock_qty":0},
    {"section":"Mitsubishi Heavy Duty Standard Inverter R32","model":"Srk10yw","btu":8683,"price_install":16800,"w_install":"1 ปี","w_parts":"5 ปี","w_comp":"5 ปี","stock_qty":0},
    {"section":"Mitsubishi Heavy Duty Standard Inverter R32","model":"Srk13yw","btu":11098,"price_install":21000,"w_install":"1 ปี","w_parts":"5 ปี","w_comp":"5 ปี","stock_qty":0},
    {"section":"Mitsubishi Heavy Duty Standard Inverter R32","model":"Srk15yw","btu":14457,"price_install":24000,"w_install":"1 ปี","w_parts":"5 ปี","w_comp":"5 ปี","stock_qty":0},
    {"section":"Mitsubishi Heavy Duty Standard Inverter R32","model":"Srk18yw","btu":17276,"price_install":28300,"w_install":"1 ปี","w_parts":"5 ปี","w_comp":"5 ปี","stock_qty":0},
    {"section":"Mitsubishi Heavy Duty Standard Inverter R32","model":"Srk24yw","btu":23021,"price_install":38000,"w_install":"1 ปี","w_parts":"5 ปี","w_comp":"5 ปี","stock_qty":0},
    {"section":"Gree Fairy Series R32","model":"Gwc09acc","btu":9000,"price_install":13500,"w_install":"1 ปี","w_parts":"5 ปี","w_comp":"10 ปี","stock_qty":0},
    {"section":"Gree Fairy Series R32","model":"Gwc12acc","btu":12000,"price_install":14700,"w_install":"1 ปี","w_parts":"5 ปี","w_comp":"10 ปี","stock_qty":0},
    {"section":"Gree Fairy Series R32","model":"Gwc18acc","btu":18000,"price_install":22500,"w_install":"1 ปี","w_parts":"5 ปี","w_comp":"10 ปี","stock_qty":0},
    {"section":"Gree Fairy Series R32","model":"Gwc24acc","btu":24000,"price_install":26500,"w_install":"1 ปี","w_parts":"5 ปี","w_comp":"10 ปี","stock_qty":0},
    {"section":"Gree Amber III R32","model":"Gwc09yb3","btu":9000,"price_install":11800,"w_install":"1 ปี","w_parts":"5 ปี","w_comp":"10 ปี","stock_qty":0},
    {"section":"Gree Amber III R32","model":"Gwc12yc3","btu":12000,"price_install":13500,"w_install":"1 ปี","w_parts":"5 ปี","w_comp":"10 ปี","stock_qty":0},
    {"section":"Gree Amber III R32","model":"Gwc18yc3","btu":18000,"price_install":20500,"w_install":"1 ปี","w_parts":"5 ปี","w_comp":"10 ปี","stock_qty":0},
    {"section":"Gree Amber III R32","model":"Gwc24yc3","btu":24000,"price_install":24000,"w_install":"1 ปี","w_parts":"5 ปี","w_comp":"10 ปี","stock_qty":0},
    {"section":"Gree Amber III Inverter R32","model":"Gwc09qb","btu":9000,"price_install":16000,"w_install":"1 ปี","w_parts":"5 ปี","w_comp":"10 ปี","stock_qty":0},
    {"section":"Gree Amber III Inverter R32","model":"Gwc12qb","btu":12000,"price_install":17000,"w_install":"1 ปี","w_parts":"5 ปี","w_comp":"10 ปี","stock_qty":0},
    {"section":"Gree Amber III Inverter R32","model":"Gwc18qb","btu":18000,"price_install":24000,"w_install":"1 ปี","w_parts":"5 ปี","w_comp":"10 ปี","stock_qty":0},
    {"section":"Gree Amber III Inverter R32","model":"Gwc24qb","btu":24000,"price_install":27300,"w_install":"1 ปี","w_parts":"5 ปี","w_comp":"10 ปี","stock_qty":0},
    {"section":"MAVELL ระบบธรรมดา","model":"MVF-09","btu":9000,"price_install":11500,"w_install":"1 ปี","w_parts":"5 ปี","w_comp":"12 ปี","stock_qty":0},
    {"section":"MAVELL ระบบธรรมดา","model":"MVF-12","btu":12000,"price_install":13000,"w_install":"1 ปี","w_parts":"5 ปี","w_comp":"12 ปี","stock_qty":0},
    {"section":"MAVELL ระบบธรรมดา","model":"MVF-18","btu":18000,"price_install":18000,"w_install":"1 ปี","w_parts":"5 ปี","w_comp":"12 ปี","stock_qty":0},
    {"section":"MAVELL ระบบธรรมดา","model":"MVF-25","btu":24000,"price_install":22500,"w_install":"1 ปี","w_parts":"5 ปี","w_comp":"12 ปี","stock_qty":0},
    {"section":"MAVELL ระบบอินเวอร์เตอร์","model":"MWF-09INV","btu":9000,"price_install":14000,"w_install":"1 ปี","w_parts":"5 ปี","w_comp":"12 ปี","stock_qty":0},
    {"section":"MAVELL ระบบอินเวอร์เตอร์","model":"MWF-12 INV","btu":12000,"price_install":15000,"w_install":"1 ปี","w_parts":"5 ปี","w_comp":"12 ปี","stock_qty":0},
    {"section":"MAVELL ระบบอินเวอร์เตอร์","model":"MWF-18 INV","btu":18000,"price_install":19800,"w_install":"1 ปี","w_parts":"5 ปี","w_comp":"12 ปี","stock_qty":0},
    {"section":"MAVELL ระบบอินเวอร์เตอร์","model":"MWF-25 INV","btu":24000,"price_install":26000,"w_install":"1 ปี","w_parts":"5 ปี","w_comp":"12 ปี","stock_qty":0},
    {"section":"DAIKIN SMASH (2018)","model":"FTM 09 PV2S","btu":9000,"price_install":14500,"w_install":"1 ปี","w_parts":"1 ปี","w_comp":"3 ปี","stock_qty":0},
    {"section":"DAIKIN SMASH (2018)","model":"FTM 13 PV2S","btu":12000,"price_install":17000,"w_install":"1 ปี","w_parts":"1 ปี","w_comp":"3 ปี","stock_qty":0},
    {"section":"DAIKIN SMASH (2018)","model":"FTM 15 PV2S","btu":15000,"price_install":20000,"w_install":"1 ปี","w_parts":"1 ปี","w_comp":"3 ปี","stock_qty":0},
    {"section":"DAIKIN SMASH (2018)","model":"FTM 18 PV2S","btu":18000,"price_install":25500,"w_install":"1 ปี","w_parts":"1 ปี","w_comp":"3 ปี","stock_qty":0},
    {"section":"DAIKIN SMASH (2018)","model":"FTM 24 PV2S","btu":24000,"price_install":35500,"w_install":"1 ปี","w_parts":"1 ปี","w_comp":"3 ปี","stock_qty":0},
    {"section":"DAIKIN SMASH (2018)","model":"FTM 28 PV2S","btu":28000,"price_install":37000,"w_install":"1 ปี","w_parts":"1 ปี","w_comp":"3 ปี","stock_qty":0},
    {"section":"DAIKIN SABAI INVERTER (2019)","model":"FTKQ 09 TV2S","btu":9000,"price_install":15500,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"5 ปี","stock_qty":0},
    {"section":"DAIKIN SABAI INVERTER (2019)","model":"FTKQ 13 TV2S","btu":12000,"price_install":18500,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"5 ปี","stock_qty":0},
    {"section":"DAIKIN SABAI INVERTER (2019)","model":"FTKQ 15 TV2S","btu":15000,"price_install":21000,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"5 ปี","stock_qty":0},
    {"section":"DAIKIN SABAI INVERTER (2019)","model":"FTKQ 18 TV2S","btu":18000,"price_install":27000,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"5 ปี","stock_qty":0},
    {"section":"DAIKIN SABAI INVERTER (2019)","model":"FTKQ 24 TV2S","btu":24000,"price_install":37000,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"5 ปี","stock_qty":0},
    {"section":"DAIKIN SUPER SMILE INVERTER","model":"FTKC 09 TV2S","btu":9000,"price_install":19000,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"5 ปี","stock_qty":0},
    {"section":"DAIKIN SUPER SMILE INVERTER","model":"FTKC 13 TV2S","btu":12000,"price_install":21000,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"5 ปี","stock_qty":0},
    {"section":"DAIKIN SUPER SMILE INVERTER","model":"FTKC 15 TV2S","btu":18000,"price_install":28500,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"5 ปี","stock_qty":0},
    {"section":"DAIKIN SUPER SMILE INVERTER","model":"FTKC 18 TV2S","btu":24000,"price_install":40500,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"5 ปี","stock_qty":0},
    {"section":"DAIKIN SUPER SMILE INVERTER","model":"FTKC 24 TV2S","btu":28000,"price_install":43500,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"5 ปี","stock_qty":0},
    {"section":"MITSUBISHI Mr.SLIM","model":"MS-GN 09 VF","btu":9000,"price_install":15500,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"5 ปี","stock_qty":0},
    {"section":"MITSUBISHI Mr.SLIM","model":"MS-GN 13 VF","btu":13000,"price_install":18500,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"5 ปี","stock_qty":0},
    {"section":"MITSUBISHI Mr.SLIM","model":"MS-GN 15 VF","btu":15000,"price_install":22500,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"5 ปี","stock_qty":0},
    {"section":"MITSUBISHI Mr.SLIM","model":"MS-GN 18 VF","btu":18000,"price_install":27000,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"5 ปี","stock_qty":0},
    {"section":"MITSUBISHI Mr.SLIM","model":"MS-GN 24 VF","btu":24000,"price_install":40000,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"5 ปี","stock_qty":0},
    {"section":"MITSUBISHI HAPPY INVERTER","model":"MSY-KP 09 VF","btu":9000,"price_install":16800,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"5 ปี","stock_qty":0},
    {"section":"MITSUBISHI HAPPY INVERTER","model":"MSY-KP 13 VF","btu":13000,"price_install":19800,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"5 ปี","stock_qty":0},
    {"section":"MITSUBISHI HAPPY INVERTER","model":"MSY-KP 15 VF","btu":15000,"price_install":23500,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"5 ปี","stock_qty":0},
    {"section":"MITSUBISHI HAPPY INVERTER","model":"MSY-KP 18 VF","btu":18000,"price_install":28500,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"5 ปี","stock_qty":0},
    {"section":"MITSUBISHI SLIM INVERTER","model":"MSY-JP 09 VF","btu":9000,"price_install":17700,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"5 ปี","stock_qty":0},
    {"section":"MITSUBISHI SLIM INVERTER","model":"MSY-JP 13 VF","btu":13000,"price_install":21000,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"5 ปี","stock_qty":0},
    {"section":"MITSUBISHI SLIM INVERTER","model":"MSY-JP 15 VF","btu":15000,"price_install":24500,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"5 ปี","stock_qty":0},
    {"section":"MITSUBISHI SLIM INVERTER","model":"MSY-JP 18 VF","btu":18000,"price_install":28500,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"5 ปี","stock_qty":0},
    {"section":"MITSUBISHI SLIM INVERTER","model":"MSY-JP 24 VF","btu":24000,"price_install":43500,"w_install":"1 ปี","w_parts":"3 ปี","w_comp":"5 ปี","stock_qty":0},
]

# ──────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────
def fmt_baht(x) -> str:
    try:
        return f"{int(x):,}"
    except Exception:
        return str(x)

def safe_text(value) -> str:
    if value is None:
        return "-"
    return str(value).replace("\r", "").strip() or "-"

def calculate_btu(w, l, h, sun, people):
    base = w * l * 900
    if h > 2.7:
        base *= 1.10
    if sun == "โดนแดด":
        base *= 1.15
    base += max(0, people - 1) * 600
    return int(round(base))

def suggest_capacity(btu):
    for s in [9000, 12000, 15000, 18000, 22000, 24000, 25000, 28000, 30000]:
        if btu <= s:
            return s
    return 30000

# ──────────────────────────────────────────────
# LOGIN
# ──────────────────────────────────────────────
def check_login():
    if "logged_in" not in st.session_state:
        st.session_state.logged_in = False
        st.session_state.username  = ""

def login_page():
    st.markdown("""
    <style>
    /* ซ่อน sidebar ตอน login */
    [data-testid="stSidebar"] { display:none !important; }
    /* login card */
    .login-wrap {
        max-width: 420px;
        margin: 40px auto 0 auto;
        background: #fff;
        border-radius: 20px;
        box-shadow: 0 4px 24px #0002;
        padding: 36px 32px 28px 32px;
        text-align: center;
    }
    .login-logo { font-size: 56px; margin-bottom: 4px; }
    .login-title { font-size: 22px; font-weight: 800; color: #0d47a1; margin: 0; }
    .login-sub   { font-size: 13px; color: #888; margin-bottom: 24px; }
    /* mobile */
    @media (max-width: 600px) {
        .login-wrap { margin: 16px 12px; padding: 28px 18px 22px 18px; }
        .login-title { font-size: 18px; }
        .login-logo  { font-size: 44px; }
    }
    </style>
    <div class="login-wrap">
        <div class="login-logo">❄️</div>
        <p class="login-title">ร้านบุญสุขอิเล็กทรอนิกส์</p>
        <p class="login-sub">Smart Sales PRO v5</p>
    </div>
    """, unsafe_allow_html=True)

    # center column — full width on mobile
    _, col, _ = st.columns([1, 3, 1])
    with col:
        st.markdown("#### 🔐 เข้าสู่ระบบ")
        username = st.text_input("👤 ชื่อผู้ใช้", placeholder="admin หรือ staff", label_visibility="visible")
        password = st.text_input("🔑 รหัสผ่าน", type="password", label_visibility="visible")
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🚀 เข้าสู่ระบบ", use_container_width=True, type="primary"):
            pw_hash = hashlib.sha256(password.encode()).hexdigest()
            if username in USERS and USERS[username] == pw_hash:
                st.session_state.logged_in = True
                st.session_state.username  = username
                st.rerun()
            else:
                st.error("❌ ชื่อผู้ใช้หรือรหัสผ่านไม่ถูกต้อง")
        st.caption("🔒 กรุณาติดต่อผู้ดูแลระบบเพื่อขอรหัสผ่าน")

# ──────────────────────────────────────────────
# SUPABASE CONNECTION
# ──────────────────────────────────────────────
def _get_supabase():
    """คืนค่า Supabase client ถ้ามี config, ไม่งั้น None"""
    try:
        from supabase import create_client
        url = st.secrets.get("SUPABASE_URL", os.environ.get("SUPABASE_URL", ""))
        key = st.secrets.get("SUPABASE_KEY", os.environ.get("SUPABASE_KEY", ""))
        if url and key and url.startswith("https://"):
            return create_client(url, key)
    except Exception:
        pass
    return None

def _use_supabase() -> bool:
    return _get_supabase() is not None

# ──────────────────────────────────────────────
# DATA LAYER  (CSV fallback → Supabase)
# ──────────────────────────────────────────────
def clean_df(df: pd.DataFrame) -> pd.DataFrame:
    defaults = {"section":"","model":"","w_install":"","w_parts":"","w_comp":"",
                "btu":0,"price_install":0,"stock_qty":0}
    for col, val in defaults.items():
        if col not in df.columns:
            df[col] = val
    df["section"]       = df["section"].astype(str).str.strip()
    df["model"]         = df["model"].astype(str).str.strip()
    df["btu"]           = pd.to_numeric(df["btu"], errors="coerce").fillna(0).astype(int)
    df["price_install"] = pd.to_numeric(df["price_install"], errors="coerce").fillna(0).astype(int)
    df["stock_qty"]     = pd.to_numeric(df["stock_qty"], errors="coerce").fillna(0).astype(int)
    df = df[(df["btu"] >= 1000) & (df["price_install"] >= 1000)]
    df = df[(df["section"] != "") & (df["model"] != "")]
    return df.copy()

@st.cache_data(ttl=60)
def load_stock() -> pd.DataFrame:
    base = clean_df(pd.DataFrame(PRODUCTS))
    # ── Supabase ──
    if _use_supabase():
        try:
            sb   = _get_supabase()
            rows = sb.table("stock").select("section,model,stock_qty").execute().data
            if rows:
                saved  = pd.DataFrame(rows)
                saved["stock_qty"] = pd.to_numeric(saved["stock_qty"], errors="coerce").fillna(0).astype(int)
                merged = base.merge(saved[["section","model","stock_qty"]], on=["section","model"],
                                    how="left", suffixes=("","_s"))
                merged["stock_qty"] = merged["stock_qty_s"].fillna(merged["stock_qty"]).astype(int)
                return merged.drop(columns=["stock_qty_s"])
        except Exception as e:
            st.warning(f"Supabase load_stock: {e} — ใช้ข้อมูลเริ่มต้น")
        return base
    # ── CSV fallback ──
    if os.path.exists(STOCK_CSV):
        try:
            saved  = clean_df(pd.read_csv(STOCK_CSV, encoding="utf-8-sig"))
            merged = base.merge(saved[["section","model","stock_qty"]], on=["section","model"],
                                how="left", suffixes=("","_s"))
            merged["stock_qty"] = merged["stock_qty_s"].fillna(merged["stock_qty"]).astype(int)
            return merged.drop(columns=["stock_qty_s"])
        except Exception:
            pass
    return base

def save_stock(df: pd.DataFrame):
    cols = ["section","model","btu","price_install","w_install","w_parts","w_comp","stock_qty"]
    # ── Supabase ──
    if _use_supabase():
        try:
            sb = _get_supabase()
            for _, r in df[cols].iterrows():
                row = r.to_dict()
                # upsert โดยใช้ section+model เป็น key
                existing = sb.table("stock").select("id").eq("section", row["section"]).eq("model", row["model"]).execute().data
                if existing:
                    sb.table("stock").update({"stock_qty": int(row["stock_qty"])}).eq("section", row["section"]).eq("model", row["model"]).execute()
                else:
                    sb.table("stock").insert({k: (int(v) if isinstance(v, (int, float)) else str(v)) for k, v in row.items()}).execute()
            st.cache_data.clear(); return
        except Exception as e:
            st.warning(f"Supabase save_stock: {e} — บันทึก CSV แทน")
    # ── CSV fallback ──
    df[cols].to_csv(STOCK_CSV, index=False, encoding="utf-8-sig")
    st.cache_data.clear()

def load_log() -> pd.DataFrame:
    # ── Supabase ──
    if _use_supabase():
        try:
            sb   = _get_supabase()
            rows = sb.table("jobs").select("*").order("created_at", desc=True).execute().data
            if rows:
                df = pd.DataFrame(rows)
                if "status"      not in df.columns: df["status"]      = JOB_STATUSES[0]
                if "receipt_no"  not in df.columns: df["receipt_no"]  = ""
                if "paid_amount" not in df.columns: df["paid_amount"] = df.get("net_total", 0)
                return df
            return pd.DataFrame()
        except Exception as e:
            st.warning(f"Supabase load_log: {e} — ใช้ CSV แทน")
    # ── CSV fallback ──
    if os.path.exists(LOG_CSV):
        try:
            df = pd.read_csv(LOG_CSV, encoding="utf-8-sig")
            if "status"      not in df.columns: df["status"]      = JOB_STATUSES[0]
            if "receipt_no"  not in df.columns: df["receipt_no"]  = ""
            if "paid_amount" not in df.columns: df["paid_amount"] = df.get("net_total", 0)
            return df
        except Exception:
            pass
    return pd.DataFrame()

def save_log(df: pd.DataFrame):
    # ── Supabase: sync ทั้ง table (update status/receipt/paid ทีละ row) ──
    if _use_supabase():
        try:
            sb = _get_supabase()
            for _, r in df.iterrows():
                row = r.to_dict()
                sb_id = row.get("id")
                update_data = {
                    "status":      str(row.get("status", "")),
                    "receipt_no":  str(row.get("receipt_no", "")),
                    "paid_amount": int(row.get("paid_amount", 0)),
                }
                if sb_id:
                    sb.table("jobs").update(update_data).eq("id", int(sb_id)).execute()
            return
        except Exception as e:
            st.warning(f"Supabase save_log: {e} — บันทึก CSV แทน")
    # ── CSV fallback ──
    df.to_csv(LOG_CSV, index=False, encoding="utf-8-sig")

def log_customer_job(quote: dict):
    record = {k: str(v).replace("\n"," | ") if isinstance(v, str) else v for k, v in quote.items()}
    record.setdefault("status",      JOB_STATUSES[0])
    record.setdefault("receipt_no",  "")
    record.setdefault("paid_amount", record.get("net_total", 0))
    record.setdefault("saved_by",    st.session_state.get("username",""))
    # ── Supabase ──
    if _use_supabase():
        try:
            sb = _get_supabase()
            insert_cols = [
                "date","customer_name","customer_phone","customer_address",
                "section","model","model_btu","base_price","discount","extra_install",
                "net_total","paid_amount","status","receipt_no","saved_by",
                "room_w","room_l","room_h","sun","people","btu","suggest_cap",
                "w_install","w_parts","w_comp"
            ]
            row = {}
            for c in insert_cols:
                val = record.get(c, "")
                if isinstance(val, float): val = int(val) if val == int(val) else val
                row[c] = val
            sb.table("jobs").insert(row).execute()
            return
        except Exception as e:
            st.warning(f"Supabase log_job: {e} — บันทึก CSV แทน")
    # ── CSV fallback ──
    pd.DataFrame([record]).to_csv(
        LOG_CSV, mode="a", header=not os.path.exists(LOG_CSV),
        index=False, encoding="utf-8-sig"
    )

def delete_job(job_id):
    """ลบงานจาก Supabase หรือ DataFrame"""
    if _use_supabase():
        try:
            sb = _get_supabase()
            sb.table("jobs").delete().eq("id", int(job_id)).execute()
            return True
        except Exception as e:
            st.warning(f"Supabase delete: {e}")
            return False
    return False

# ──────────────────────────────────────────────
# EXCEL EXPORT
# ──────────────────────────────────────────────
def export_excel(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:  # requires: pip install openpyxl
        cols_job = [c for c in [
            "date","receipt_no","customer_name","customer_phone",
            "section","model","model_btu","base_price","discount",
            "extra_install","net_total","paid_amount","status","saved_by"
        ] if c in df.columns]
        df_out = df[cols_job].copy()
        rename = {
            "date":"วันที่","receipt_no":"เลขใบเสร็จ","customer_name":"ชื่อลูกค้า",
            "customer_phone":"เบอร์โทร","section":"ซีรีส์","model":"รุ่น",
            "model_btu":"BTU","base_price":"ราคาพร้อมติดตั้ง","discount":"ส่วนลด",
            "extra_install":"ติดตั้งเพิ่ม","net_total":"รวมสุทธิ",
            "paid_amount":"รับเงินแล้ว","status":"สถานะ","saved_by":"บันทึกโดย"
        }
        df_out.rename(columns=rename, inplace=True)
        df_out.to_excel(writer, sheet_name="รายการงาน", index=False)

        # Sheet สรุป
        if "net_total" in df.columns:
            df2 = df.copy()
            df2["net_total"]   = pd.to_numeric(df2["net_total"],   errors="coerce").fillna(0)
            df2["paid_amount"] = pd.to_numeric(df2.get("paid_amount", df2["net_total"]), errors="coerce").fillna(0)
            paid_count = len(df2[df2["status"] == "💰 รับเงินแล้ว"]) if "status" in df2.columns else 0
            wait_count = len(df2[df2["status"] == "📋 รอดำเนินการ"]) if "status" in df2.columns else 0
            summary = pd.DataFrame({
                "รายการ": ["จำนวนงานทั้งหมด","ยอดขายรวม (บาท)","รับเงินแล้ว (บาท)","เฉลี่ย/งาน (บาท)","งานรับเงินแล้ว","งานรอดำเนินการ"],
                "ค่า":    [len(df2), df2["net_total"].sum(), df2["paid_amount"].sum(),
                            round(df2["net_total"].mean(),2) if len(df2) else 0, paid_count, wait_count]
            })
            summary.to_excel(writer, sheet_name="สรุปยอดขาย", index=False)

        wb = writer.book
        from openpyxl.styles import Font, PatternFill, Alignment
        hfill = PatternFill("solid", fgColor="1565C0")
        hfont = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.fill = hfill; cell.font = hfont
                cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                mlen = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(mlen + 4, 40)
    return buf.getvalue()

# ──────────────────────────────────────────────
# PDF HELPERS
# ──────────────────────────────────────────────
def _load_thai_font(pdf: FPDF) -> str:
    os.makedirs("fonts", exist_ok=True)
    zip_path = "THSarabunNew.zip"; extract_dir = "fonts"
    if os.path.exists(zip_path):
        has_ttf = any(f.lower().endswith(".ttf") for _,_,files in os.walk(extract_dir) for f in files)
        if not has_ttf:
            with zipfile.ZipFile(zip_path,"r") as z: z.extractall(extract_dir)
    regular, bold = [], []
    for root, _, files in os.walk(extract_dir):
        for f in files:
            if f.lower().endswith(".ttf") and ("thsarabunnew" in f.lower() or "sarabun" in f.lower()):
                path = os.path.join(root, f)
                (bold if "bold" in f.lower() else regular).append(path)
    if not regular: raise FileNotFoundError("ไม่พบฟอนต์ภาษาไทย .ttf")
    pdf.add_font("THSarabun","",  regular[0], uni=True)
    pdf.add_font("THSarabun","B", (bold or regular)[0], uni=True)
    return "THSarabun"

def _pdf_header(pdf, font, CW, L, title):
    def go(): pdf.set_x(L)
    logo = os.path.join("assets","logo.png")
    if os.path.exists(logo):
        try: pdf.image(logo, x=L, y=10, w=22)
        except: pass
    pdf.set_font(font,"B",20); go(); pdf.cell(CW,8,title,ln=1,align="C")
    pdf.set_font(font,"B",14); go(); pdf.cell(CW,7,STORE_NAME,ln=1,align="C")
    pdf.set_font(font,"",11)
    go(); pdf.cell(CW,6,f"ที่อยู่: {STORE_ADDRESS}",ln=1,align="C")
    go(); pdf.cell(CW,6,f"มือถือ/Line: {STORE_PHONE}  |  {STORE_WEB}",ln=1,align="C")
    if STORE_TAX_ID:
        go(); pdf.cell(CW,6,f"เลขผู้เสียภาษี: {STORE_TAX_ID}",ln=1,align="C")
    pdf.ln(2); pdf.line(L,pdf.get_y(),L+CW,pdf.get_y()); pdf.ln(3)

def _pdf_out(pdf):
    out = pdf.output(dest="S")
    return bytes(out) if isinstance(out,(bytes,bytearray)) else out.encode("latin-1")

def _money_row(pdf, font, CW, L, label, value, bold=False):
    pdf.set_x(L)
    pdf.set_font(font,"B" if bold else "",12)
    pdf.cell(130,7,label)
    pdf.cell(CW-130,7,f"{fmt_baht(value)} บาท",ln=1,align="R")

def _signatures(pdf, font, CW, L, left_label, right_label):
    pdf.ln(5); pdf.set_font(font,"",11)
    col = CW//2-3; gap = CW-col*2
    pdf.set_x(L)
    pdf.cell(col,8,f"{left_label} .................................................")
    pdf.cell(gap,8,"")
    pdf.cell(col,8,f"{right_label} .................................................",ln=1)
    pdf.set_x(L)
    pdf.cell(col,8,f"({STORE_NAME})",align="C")
    pdf.cell(gap,8,"")
    pdf.cell(col,8,"(..................................)",ln=1,align="C")
    pdf.set_x(L)
    pdf.cell(col,8,"วันที่ ........../........../..........",align="C")
    pdf.cell(gap,8,"")
    pdf.cell(col,8,"วันที่ ........../........../..........",ln=1,align="C")

# ── ใบเสนอราคา ───────────────────────────────
def build_pdf_quotation(q: dict) -> bytes:
    pdf = FPDF(unit="mm",format="A4"); pdf.add_page()
    pdf.set_auto_page_break(auto=True,margin=12)
    font = _load_thai_font(pdf); L=15; CW=180
    _pdf_header(pdf,font,CW,L,"ใบเสนอราคา / QUOTATION")

    def go(): pdf.set_x(L)
    def row(lbl,val,lw=35):
        go(); pdf.set_font(font,"B",12); pdf.cell(lw,6,lbl)
        pdf.set_font(font,"",12); pdf.multi_cell(CW-lw,6,safe_text(val))
    def sec(t):
        pdf.ln(2); go(); pdf.set_font(font,"B",13); pdf.cell(CW,7,t,ln=1)

    doc_no = f"QT-{datetime.today().strftime('%Y%m%d%H%M')}"
    go()
    pdf.set_font(font,"B",12); pdf.cell(25,7,"วันที่")
    pdf.set_font(font,"",12);  pdf.cell(60,7,safe_text(q["date"]))
    pdf.set_font(font,"B",12); pdf.cell(35,7,"เลขที่เอกสาร")
    pdf.set_font(font,"",12);  pdf.cell(60,7,doc_no,ln=1)

    sec("ข้อมูลลูกค้า")
    row("ชื่อลูกค้า", q["customer_name"])
    row("เบอร์โทร",  q["customer_phone"])
    row("ที่อยู่",   q["customer_address"])

    sec("รายการสินค้า")
    top=pdf.get_y(); bh=38; pdf.rect(L,top,CW,bh)
    pdf.set_xy(L+3,top+3); pdf.set_font(font,"B",12)
    pdf.multi_cell(CW-6,6,safe_text(q["section"]))
    pdf.set_x(L+3); pdf.set_font(font,"",12)
    pdf.multi_cell(CW-6,6,f"Model: {safe_text(q['model'])}   |   BTU: {q['model_btu']:,}")
    pdf.set_x(L+3)
    pdf.multi_cell(CW-6,6,f"ประกัน: ติดตั้ง {safe_text(q['w_install'])} / อะไหล่ {safe_text(q['w_parts'])} / คอมฯ {safe_text(q['w_comp'])}")
    if pdf.get_y() < top+bh: pdf.set_y(top+bh)

    sec("สรุปราคา")
    sy=pdf.get_y(); sh=34; pdf.rect(L,sy,CW,sh)
    pdf.set_xy(L+4,sy+4)
    _money_row(pdf,font,CW,L+4,"ราคาพร้อมติดตั้ง",q["base_price"])
    _money_row(pdf,font,CW,L+4,"ส่วนลด",-int(q["discount"]))
    _money_row(pdf,font,CW,L+4,"ค่าติดตั้งเพิ่ม",int(q["extra_install"]))
    pdf.line(L+4,pdf.get_y(),L+CW-4,pdf.get_y()); pdf.ln(2)
    _money_row(pdf,font,CW,L+4,"รวมสุทธิ",q["net_total"],bold=True)
    if pdf.get_y() < sy+sh: pdf.set_y(sy+sh)

    sec("เงื่อนไขการติดตั้ง")
    pdf.set_font(font,"",11)
    for cline in INSTALL_CONDITIONS.split("\n"):
        if cline.strip(): go(); pdf.multi_cell(CW,5.5,cline.strip())

    _signatures(pdf,font,CW,L,"ลงชื่อผู้เสนอราคา","ลงชื่อผู้รับใบเสนอราคา")
    return _pdf_out(pdf)

# ── ใบเสร็จ / ใบกำกับภาษี ───────────────────
def build_pdf_receipt(q: dict, receipt_no: str, is_tax: bool = False) -> bytes:
    title = "ใบกำกับภาษีอย่างง่าย" if is_tax else "ใบเสร็จรับเงิน"
    pdf = FPDF(unit="mm",format="A4"); pdf.add_page()
    pdf.set_auto_page_break(auto=True,margin=12)
    font = _load_thai_font(pdf); L=15; CW=180
    _pdf_header(pdf,font,CW,L,title)

    def go(): pdf.set_x(L)
    def row(lbl,val,lw=40):
        go(); pdf.set_font(font,"B",12); pdf.cell(lw,6,lbl)
        pdf.set_font(font,"",12); pdf.multi_cell(CW-lw,6,safe_text(val))

    # doc ref
    go()
    pdf.set_font(font,"B",12); pdf.cell(40,7,"เลขที่ใบเสร็จ")
    pdf.set_font(font,"",12);  pdf.cell(60,7,receipt_no)
    pdf.set_font(font,"B",12); pdf.cell(20,7,"วันที่")
    pdf.set_font(font,"",12);  pdf.cell(60,7,safe_text(q["date"]),ln=1)
    pdf.ln(2)

    pdf.set_font(font,"B",13); go(); pdf.cell(CW,7,"ข้อมูลลูกค้า",ln=1)
    row("ชื่อ",   q.get("customer_name",""))
    row("โทร",    q.get("customer_phone",""))
    row("ที่อยู่", q.get("customer_address",""))
    if is_tax:
        row("เลขผู้เสียภาษี (ลูกค้า)", q.get("customer_tax_id","-"))

    # item table
    pdf.ln(3); pdf.set_font(font,"B",12); go()
    pdf.cell(10,7,"ที่",border=1,align="C")
    pdf.cell(90,7,"รายการ",border=1,align="C")
    pdf.cell(30,7,"จำนวน",border=1,align="C")
    pdf.cell(50,7,"ราคา (บาท)",border=1,align="C",ln=1)
    pdf.set_font(font,"",12); go()
    item_desc = f"{safe_text(q.get('section',''))} {safe_text(q.get('model',''))} {int(q.get('model_btu',0)):,} BTU"
    pdf.cell(10,7,"1",border=1,align="C")
    pdf.cell(90,7,item_desc[:45],border=1)
    pdf.cell(30,7,"1 เครื่อง",border=1,align="C")
    pdf.cell(50,7,fmt_baht(q.get("base_price",0)),border=1,align="R",ln=1)

    pdf.ln(3)
    if int(q.get("discount",0)) > 0:
        _money_row(pdf,font,CW,L,"ส่วนลด",-int(q["discount"]))
    if int(q.get("extra_install",0)) > 0:
        _money_row(pdf,font,CW,L,"ค่าติดตั้งเพิ่ม",int(q["extra_install"]))

    if is_tax:
        before_vat = round(q["net_total"]/1.07,2)
        vat = q["net_total"]-before_vat
        go(); pdf.line(L,pdf.get_y(),L+CW,pdf.get_y()); pdf.ln(1)
        _money_row(pdf,font,CW,L,"ราคาก่อน VAT 7%",before_vat)
        _money_row(pdf,font,CW,L,"VAT 7%",vat)

    go(); pdf.line(L,pdf.get_y(),L+CW,pdf.get_y()); pdf.ln(2)
    _money_row(pdf,font,CW,L,"รวมสุทธิ",q["net_total"],bold=True)
    pdf.ln(2); go()
    pdf.set_font(font,"B",13)
    paid = q.get("paid_amount", q.get("net_total",0))
    pdf.cell(CW,8,f"✅ รับชำระเงินแล้ว: {fmt_baht(paid)} บาท",ln=1,align="C")

    _signatures(pdf,font,CW,L,"ลงชื่อผู้รับเงิน","ลงชื่อผู้จ่ายเงิน")
    return _pdf_out(pdf)

# ──────────────────────────────────────────────
# LINE
# ──────────────────────────────────────────────
def make_line_text(q: dict) -> str:
    return "\n".join([
        f"ใบเสนอราคา - {STORE_NAME}",
        f"วันที่: {q['date']}",
        f"ลูกค้า: {safe_text(q['customer_name'])}",
        f"โทร: {safe_text(q['customer_phone'])}",
        f"รุ่น: {safe_text(q['model'])}  ({safe_text(q['section'])})",
        f"BTU รุ่น: {q['model_btu']:,} BTU",
        f"ราคาพร้อมติดตั้ง: {fmt_baht(q['base_price'])} บาท",
        f"ส่วนลด: {fmt_baht(q['discount'])} บาท",
        f"ค่าติดตั้งเพิ่ม: {fmt_baht(q['extra_install'])} บาท",
        f"รวมสุทธิ: {fmt_baht(q['net_total'])} บาท",
        "",
        f"ติดต่อร้าน: {STORE_PHONE}",
        STORE_WEB,
    ])

def line_share_link(text): return "https://line.me/R/msg/text/?" + urlquote(text)

# ──────────────────────────────────────────────
# CSS
# ──────────────────────────────────────────────
st.markdown("""
<style>
/* ── Sidebar ─────────────────────────────── */
[data-testid="stSidebar"] { background:linear-gradient(180deg,#0d47a1 0%,#1565c0 100%); }
[data-testid="stSidebar"] * { color:#fff !important; }

/* ── Metric cards ────────────────────────── */
.metric-card { background:#f0f4ff;border-radius:12px;padding:16px 20px;border-left:5px solid #1565c0;margin-bottom:10px; }
.metric-card h4 { margin:0;color:#1565c0;font-size:13px; }
.metric-card h2 { margin:4px 0 0;color:#0d47a1;font-size:26px;font-weight:800; }

/* ── Stock badges ────────────────────────── */
.badge-in  { background:#e8f5e9;color:#2e7d32;padding:2px 10px;border-radius:20px;font-weight:700; }
.badge-out { background:#fce4ec;color:#c62828;padding:2px 10px;border-radius:20px;font-weight:700; }
.badge-low { background:#fff3e0;color:#e65100;padding:2px 10px;border-radius:20px;font-weight:700; }

/* ══ MOBILE RESPONSIVE (max-width: 768px) ═══ */
@media (max-width: 768px) {

  /* ขยาย content เต็มจอ */
  .main .block-container {
    padding: 0.5rem 0.8rem 2rem !important;
    max-width: 100% !important;
  }

  /* ปุ่มใหญ่ขึ้น กดง่าย */
  .stButton > button {
    height: 3rem !important;
    font-size: 16px !important;
    border-radius: 10px !important;
    width: 100% !important;
  }

  /* Input box ใหญ่ขึ้น */
  .stTextInput input, .stSelectbox select,
  .stNumberInput input, .stTextArea textarea {
    font-size: 16px !important;
    height: 2.8rem !important;
    border-radius: 8px !important;
  }

  /* หัวข้อเล็กลงนิดหน่อย */
  h1 { font-size: 1.4rem !important; }
  h2 { font-size: 1.2rem !important; }
  h3 { font-size: 1.1rem !important; }

  /* Expander ขยาย padding */
  .streamlit-expanderHeader {
    font-size: 15px !important;
    padding: 12px !important;
  }

  /* Columns บนมือถือเรียงแนวตั้ง */
  [data-testid="column"] {
    width: 100% !important;
    flex: 1 1 100% !important;
    min-width: 100% !important;
  }

  /* Tab text ขนาดพอดี */
  .stTabs [data-baseweb="tab"] {
    font-size: 13px !important;
    padding: 8px 10px !important;
  }

  /* Dataframe scroll ได้ */
  .stDataFrame {
    overflow-x: auto !important;
    font-size: 12px !important;
  }

  /* Metric card บนมือถือ */
  .metric-card h2 { font-size: 20px !important; }
  .metric-card h4 { font-size: 12px !important; }

  /* Download button */
  .stDownloadButton > button {
    height: 3rem !important;
    font-size: 15px !important;
    width: 100% !important;
  }

  /* Warning / info box */
  .stAlert { font-size: 14px !important; border-radius: 10px !important; }

  /* Sidebar toggle button ใหญ่ขึ้น */
  [data-testid="collapsedControl"] {
    width: 2.5rem !important;
    height: 2.5rem !important;
  }
}
</style>""", unsafe_allow_html=True)

# ──────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────
check_login()
if not st.session_state.logged_in:
    login_page(); st.stop()

# Sidebar
with st.sidebar:
    logo_path = os.path.join("assets","logo.png")
    if os.path.exists(logo_path): st.image(logo_path, use_container_width=True)
    st.markdown(f"## ❄️ {STORE_NAME}")
    st.markdown(f"📍 {STORE_ADDRESS}")
    st.markdown(f"📞 **{STORE_PHONE}**")
    st.markdown(f"🌐 [Facebook]({STORE_WEB})")
    st.divider()
    page = st.radio("เมนู", [
        "🧾 สร้างใบเสนอราคา",
        "🛠️ รับงานซ่อม/บริการ",
        "📋 จัดการงาน / สถานะ",
        "📦 จัดการสต๊อก",
        "📊 Dashboard",
        "🔧 คลังเออเร่อแอร์",
    ] + (["⚙️ นำเข้า/ส่งออกข้อมูล"] if st.session_state.get("username") == "admin" else []),
    label_visibility="collapsed")
    st.divider()
    st.caption(f"👤 ล็อกอิน: **{st.session_state.username}**")
    if st.button("🚪 ออกจากระบบ", use_container_width=True):
        st.session_state.logged_in = False; st.session_state.username = ""; st.rerun()

df_all = load_stock()

# แจ้งเตือนสต๊อกใกล้หมด
low_stock = df_all[df_all["stock_qty"] <= LOW_STOCK_THRESHOLD]
if not low_stock.empty:
    items = ", ".join(f"{r['model']} (เหลือ {r['stock_qty']})" for _, r in low_stock.head(5).iterrows())
    st.warning(f"⚠️ **สต๊อกใกล้หมด:** {items}")

# ══════════════════════════════════════════════
# PAGE 1: QUOTATION
# ══════════════════════════════════════════════
if page == "🧾 สร้างใบเสนอราคา":
    st.title("🧾 สร้างใบเสนอราคา")

    with st.expander("👤 ขั้นตอน 1 — ข้อมูลลูกค้า", expanded=True):
        c1, c2 = st.columns(2)
        customer_name    = c1.text_input("ชื่อลูกค้า")
        customer_phone   = c2.text_input("เบอร์โทร")
        customer_address = st.text_area("ที่อยู่/สถานที่ติดตั้ง", height=68)

    with st.expander("🏠 ขั้นตอน 2 — ขนาดห้องและ BTU", expanded=True):
        c1, c2, c3 = st.columns(3)
        room_w = c1.number_input("กว้าง (เมตร)", min_value=0.0, step=0.1)
        room_l = c2.number_input("ยาว (เมตร)",   min_value=0.0, step=0.1)
        room_h = c3.number_input("สูง (เมตร)",   min_value=0.0, step=0.1, value=2.6)
        c4, c5 = st.columns(2)
        sun    = c4.selectbox("ห้องโดนแดด?", ["ไม่โดนแดด","โดนแดด"])
        people = c5.number_input("จำนวนคน", min_value=1, step=1, value=1)
        btu = suggest_cap = None
        if room_w > 0 and room_l > 0:
            btu = calculate_btu(room_w, room_l, room_h, sun, int(people))
            suggest_cap = suggest_capacity(btu)
            m1, m2 = st.columns(2)
            m1.metric("BTU ที่คำนวณได้", f"{btu:,} BTU")
            m2.metric("ขนาดแอร์แนะนำ",  f"{suggest_cap:,} BTU")
        else:
            st.info("กรอกขนาดห้องเพื่อคำนวณ BTU")

    with st.expander("❄️ ขั้นตอน 3 — เลือกรุ่นแอร์", expanded=True):
        q_search = st.text_input("🔍 ค้นหา", placeholder="เช่น Fujitsu, 12000, inverter").strip().lower()
        df_view  = df_all.copy()
        if q_search:
            mask = (df_view["section"].str.lower().str.contains(q_search,na=False) |
                    df_view["model"].str.lower().str.contains(q_search,na=False))
            df_view = df_view[mask]
        if btu and suggest_cap:
            if st.checkbox(f"แสดงเฉพาะรุ่น BTU ใกล้เคียง ({suggest_cap:,} BTU)", value=False):
                df_view = df_view[(df_view["btu"] >= suggest_cap*0.8) & (df_view["btu"] <= suggest_cap*1.3)]

        sections = sorted(df_view["section"].dropna().unique().tolist())
        if not sections: st.warning("ไม่พบสินค้า"); st.stop()

        col_s, col_m = st.columns([2,2])
        section  = col_s.selectbox("ซีรีส์/หมวดรุ่น", options=sections)
        df_sec   = df_view[df_view["section"] == section].copy()
        if btu and suggest_cap:
            df_sec["_diff"] = (df_sec["btu"]-suggest_cap).abs()
            df_sec = df_sec.sort_values(["_diff","price_install"])
        else:
            df_sec = df_sec.sort_values("price_install")
        model    = col_m.selectbox("Model", options=df_sec["model"].tolist())
        row_data = df_sec[df_sec["model"]==model].iloc[0].to_dict()
        sq = int(row_data.get("stock_qty",0))
        badge = (f'<span class="badge-in">มีสต๊อก {sq} เครื่อง</span>' if sq > LOW_STOCK_THRESHOLD
                 else f'<span class="badge-low">เหลือ {sq} เครื่อง</span>' if sq > 0
                 else '<span class="badge-out">สต๊อกหมด</span>')
        ca, cb, cc, cd = st.columns(4)
        ca.metric("BTU",f"{int(row_data['btu']):,}")
        cb.metric("ราคา (พร้อมติดตั้ง)",f"{fmt_baht(row_data['price_install'])} ฿")
        cc.metric("ประกันคอมฯ",row_data.get("w_comp","-"))
        cd.markdown(f"**สต๊อก**<br>{badge}", unsafe_allow_html=True)
        st.caption(f"ประกันติดตั้ง {safe_text(row_data.get('w_install'))} | ประกันอะไหล่ {safe_text(row_data.get('w_parts'))}")

    with st.expander("💰 ขั้นตอน 4 — ปรับราคา", expanded=True):
        p1, p2 = st.columns(2)
        discount      = p1.number_input("ส่วนลด (บาท)", min_value=0, step=100, value=0)
        extra_install = p2.number_input("ค่าติดตั้งเพิ่ม (บาท)", min_value=0, step=100, value=0)
        base_price = int(row_data["price_install"])
        net_total  = max(0, base_price - int(discount) + int(extra_install))
        st.markdown(f"""<div class="metric-card"><h4>สรุปราคา</h4><h2>฿ {fmt_baht(net_total)}</h2>
        <small>ราคาพร้อมติดตั้ง {fmt_baht(base_price)} | ส่วนลด {fmt_baht(discount)} | ติดตั้งเพิ่ม {fmt_baht(extra_install)}</small>
        </div>""", unsafe_allow_html=True)

    today_str  = date.today().strftime("%d/%m/%Y")
    quote_data = dict(
        date=today_str, customer_name=customer_name, customer_phone=customer_phone,
        customer_address=customer_address, room_w=room_w, room_l=room_l, room_h=room_h,
        sun=sun, people=int(people),
        btu=int(btu) if btu else 0, suggest_cap=int(suggest_cap) if suggest_cap else 0,
        section=section, model=model, model_btu=int(row_data["btu"]),
        w_install=row_data.get("w_install",""), w_parts=row_data.get("w_parts",""),
        w_comp=row_data.get("w_comp",""),
        base_price=base_price, discount=int(discount),
        extra_install=int(extra_install), net_total=int(net_total),
        status=JOB_STATUSES[0], saved_by=st.session_state.username,
    )

    st.divider(); st.subheader("📤 ดำเนินการ")
    a1, a2, a3 = st.columns(3)
    if a1.button("💾 บันทึกงาน", use_container_width=True, type="primary"):
        log_customer_job(quote_data); st.success("บันทึกแล้ว ✅")
    if a2.button("📄 สร้าง PDF ใบเสนอราคา", use_container_width=True):
        try:
            pdf_b = build_pdf_quotation(quote_data)
            fname = f"ใบเสนอราคา_{customer_name or 'ลูกค้า'}_{today_str.replace('/','')}.pdf"
            st.success("สร้าง PDF สำเร็จ ✅")
            st.download_button("⬇️ ดาวน์โหลด PDF", data=pdf_b, file_name=fname,
                               mime="application/pdf", use_container_width=True)
        except Exception as e:
            st.error(f"สร้าง PDF ไม่สำเร็จ: {e}")
    line_text = make_line_text(quote_data)
    a3.markdown(f"""<a href="{line_share_link(line_text)}" target="_blank" style="
        display:block;text-align:center;background:#00c300;color:white;
        padding:10px;border-radius:8px;text-decoration:none;font-weight:700;font-size:15px;margin-top:4px;">
        💬 ส่ง LINE</a>""", unsafe_allow_html=True)
    with st.expander("📝 ดูข้อความ LINE"):
        st.text_area("ข้อความ", value=line_text, height=180)

# ══════════════════════════════════════════════
# PAGE 2: JOB MANAGEMENT
# ══════════════════════════════════════════════
elif page == "📋 จัดการงาน / สถานะ":
    st.title("📋 จัดการงานและสถานะ")

    tab_ac, tab_sv = st.tabs(["❄️ งานแอร์ / ใบเสนอราคา", "🛠️ งานซ่อม / บริการ"])

    # ══ Tab 1: งานแอร์ ══════════════════════════
    with tab_ac:
        df_log = load_log()
        if df_log.empty:
            st.info("ยังไม่มีงานแอร์ กรุณาบันทึกจากหน้าใบเสนอราคาก่อน")
        else:
            f1, f2, f3 = st.columns(3)
            filter_status = f1.selectbox("กรองตามสถานะ", ["ทั้งหมด"] + JOB_STATUSES, key="ac_fs")
            filter_name   = f2.text_input("ค้นหาชื่อ/เบอร์", key="ac_fn")
            filter_date   = f3.text_input("ค้นหาวันที่", key="ac_fd")

            df_show = df_log.copy()
            if filter_status != "ทั้งหมด" and "status" in df_show.columns:
                df_show = df_show[df_show["status"] == filter_status]
            if filter_name:
                mask = pd.Series(False, index=df_show.index)
                for col in ["customer_name","customer_phone"]:
                    if col in df_show.columns:
                        mask |= df_show[col].astype(str).str.lower().str.contains(filter_name.lower(),na=False)
                df_show = df_show[mask]
            if filter_date and "date" in df_show.columns:
                df_show = df_show[df_show["date"].astype(str).str.contains(filter_date,na=False)]

            st.markdown(f"**พบ {len(df_show)} รายการ**")

            for idx, job in df_show.iterrows():
                status = job.get("status", JOB_STATUSES[0])
                icon   = "💰" if "รับเงิน" in status else "✅" if "ติดตั้งแล้ว" in status else "🔧" if "กำลัง" in status else "❌" if "ยกเลิก" in status else "📋"
                with st.expander(f"{icon} {job.get('customer_name','?')} | {job.get('model','?')} | {job.get('date','?')} | ฿{fmt_baht(job.get('net_total',0))} | {status}"):
                    c1, c2 = st.columns(2)
                    c1.markdown(f"**ลูกค้า:** {job.get('customer_name','-')}")
                    c1.markdown(f"**โทร:** {job.get('customer_phone','-')}")
                    c1.markdown(f"**รุ่น:** {job.get('model','-')} | {int(job.get('model_btu',0)):,} BTU")
                    c2.markdown(f"**ราคาสุทธิ:** ฿{fmt_baht(job.get('net_total',0))}")
                    c2.markdown(f"**รับเงิน:** ฿{fmt_baht(job.get('paid_amount',job.get('net_total',0)))}")
                    c2.markdown(f"**บันทึกโดย:** {job.get('saved_by','-')}")
                    st.divider()
                    e1, e2, e3 = st.columns(3)
                    new_status  = e1.selectbox("เปลี่ยนสถานะ", JOB_STATUSES,
                                    index=JOB_STATUSES.index(status) if status in JOB_STATUSES else 0,
                                    key=f"st_{idx}")
                    receipt_no  = e2.text_input("เลขที่ใบเสร็จ", value=str(job.get("receipt_no","")), key=f"rn_{idx}")
                    paid_amount = e3.number_input("จำนวนเงินที่รับ (฿)", value=int(job.get("paid_amount",job.get("net_total",0))), step=100, key=f"pa_{idx}")

                    u1, u2, u3, u4 = st.columns(4)
                    if u1.button("💾 อัปเดต", key=f"upd_{idx}", use_container_width=True, type="primary"):
                        df_log.at[idx,"status"]      = new_status
                        df_log.at[idx,"receipt_no"]  = receipt_no
                        df_log.at[idx,"paid_amount"] = paid_amount
                        save_log(df_log); st.success("อัปเดตแล้ว ✅"); st.rerun()

                    if u2.button("🧾 ใบเสร็จ", key=f"rc_{idx}", use_container_width=True):
                        try:
                            qd = job.to_dict(); qd["paid_amount"] = paid_amount
                            pb = build_pdf_receipt(qd, receipt_no or f"RC-{idx}", is_tax=False)
                            st.download_button("⬇️ ดาวน์โหลดใบเสร็จ", data=pb,
                                               file_name=f"ใบเสร็จ_{receipt_no or idx}.pdf",
                                               mime="application/pdf", key=f"dlrc_{idx}")
                        except Exception as e: st.error(f"ไม่สำเร็จ: {e}")

                    if u3.button("📑 ใบกำกับภาษี", key=f"tx_{idx}", use_container_width=True):
                        try:
                            qd = job.to_dict(); qd["paid_amount"] = paid_amount
                            pb = build_pdf_receipt(qd, receipt_no or f"TX-{idx}", is_tax=True)
                            st.download_button("⬇️ ดาวน์โหลดใบกำกับภาษี", data=pb,
                                               file_name=f"ใบกำกับภาษี_{receipt_no or idx}.pdf",
                                               mime="application/pdf", key=f"dltx_{idx}")
                        except Exception as e: st.error(f"ไม่สำเร็จ: {e}")

                    if u4.button("🗑️ ลบ", key=f"del_{idx}", use_container_width=True):
                        df_log = df_log.drop(index=idx).reset_index(drop=True)
                        save_log(df_log); st.warning("ลบแล้ว"); st.rerun()

            st.divider()
            if st.button("📊 Export Excel รายงานแอร์", use_container_width=True, type="primary"):
                xlsx = export_excel(df_log)
                st.download_button("⬇️ ดาวน์โหลด Excel", data=xlsx,
                                   file_name=f"รายงาน_บุญสุข_{date.today().strftime('%Y%m%d')}.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   use_container_width=True)

    # ══ Tab 2: งานซ่อม/บริการ ═══════════════════
    with tab_sv:
        df_sv2 = load_service()
        if df_sv2.empty:
            st.info("ยังไม่มีงานซ่อม/บริการ กรุณาบันทึกจากหน้ารับงานซ่อม/บริการก่อน")
        else:
            sf1, sf2, sf3 = st.columns(3)
            sv2_filter_status = sf1.selectbox("กรองตามสถานะ", ["ทั้งหมด"] + SERVICE_STATUSES, key="sv2_fs")
            sv2_filter_type   = sf2.selectbox("กรองประเภทงาน", ["ทั้งหมด"] + SERVICE_TYPES, key="sv2_ft")
            sv2_search        = sf3.text_input("ค้นหาชื่อ/เบอร์", key="sv2_srch")

            df_sv2_show = df_sv2.copy()
            if sv2_filter_status != "ทั้งหมด":
                df_sv2_show = df_sv2_show[df_sv2_show["status"] == sv2_filter_status]
            if sv2_filter_type != "ทั้งหมด":
                df_sv2_show = df_sv2_show[df_sv2_show["service_type"] == sv2_filter_type]
            if sv2_search:
                mask2 = (df_sv2_show["customer_name"].astype(str).str.contains(sv2_search, case=False, na=False) |
                         df_sv2_show["customer_phone"].astype(str).str.contains(sv2_search, case=False, na=False))
                df_sv2_show = df_sv2_show[mask2]

            st.markdown(f"**พบ {len(df_sv2_show)} รายการ**")

            for idx2, r2 in df_sv2_show.iterrows():
                st2 = str(r2.get("status", SERVICE_STATUSES[0]))
                icon2 = "💰" if "รับเงิน" in st2 else "✅" if "เสร็จ" in st2 else "🔧" if "ดำเนินการ" in st2 else "❌" if "ยกเลิก" in st2 else "📋"
                with st.expander(
                    f"{icon2} {r2.get('service_type','')} — {r2.get('customer_name','?')} | {r2.get('customer_phone','?')} | {r2.get('date','?')} | ฿{fmt_baht(r2.get('price',0))} | {st2}"
                ):
                    d1, d2 = st.columns(2)
                    d1.markdown(f"**👤 ลูกค้า:** {r2.get('customer_name','-')}")
                    d1.markdown(f"**📞 โทร:** {r2.get('customer_phone','-')}")
                    d1.markdown(f"**📍 ที่อยู่:** {r2.get('customer_address','-')}")
                    d2.markdown(f"**💰 ค่าบริการ:** ฿{fmt_baht(r2.get('price',0))}")
                    d2.markdown(f"**📅 วันที่:** {r2.get('date','-')}")
                    d2.markdown(f"**บันทึกโดย:** {r2.get('saved_by','-')}")
                    st.markdown(f"**⚡ อาการ/งาน:** {r2.get('symptom','-')}")
                    if r2.get("note",""):
                        st.markdown(f"**📌 หมายเหตุ:** {r2.get('note','')}")
                    st.divider()

                    sv2_u1, sv2_u2, sv2_u3 = st.columns([3,2,1])
                    new_st2 = sv2_u1.selectbox("เปลี่ยนสถานะ", SERVICE_STATUSES,
                                index=SERVICE_STATUSES.index(st2) if st2 in SERVICE_STATUSES else 0,
                                key=f"sv2_st_{idx2}")
                    new_price2 = sv2_u2.number_input("ค่าบริการ (฿)", min_value=0, step=50,
                                value=int(r2.get("price",0)), key=f"sv2_pr_{idx2}")
                    if sv2_u3.button("💾", key=f"sv2_save_{idx2}", help="บันทึก"):
                        df_sv2.at[idx2,"status"] = new_st2
                        df_sv2.at[idx2,"price"]  = new_price2
                        save_service(df_sv2)
                        st.success("✅ อัปเดตแล้ว"); st.rerun()

                    if st.button("🗑️ ลบงานนี้", key=f"sv2_del_{idx2}", use_container_width=True):
                        df_sv2 = df_sv2.drop(index=idx2).reset_index(drop=True)
                        save_service(df_sv2); st.warning("ลบแล้ว"); st.rerun()

            st.divider()
            csv_sv2 = df_sv2_show.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")
            st.download_button("⬇️ Export CSV งานซ่อม/บริการ", data=csv_sv2,
                               file_name=f"service_jobs_{date.today().strftime('%Y%m%d')}.csv",
                               mime="text/csv", use_container_width=True)

# ══════════════════════════════════════════════
# PAGE 3: STOCK
# ══════════════════════════════════════════════
elif page == "📦 จัดการสต๊อก":
    st.title("📦 จัดการสต๊อกแอร์")
    total_m = len(df_all)
    in_s    = len(df_all[df_all["stock_qty"] > LOW_STOCK_THRESHOLD])
    low_s   = len(df_all[(df_all["stock_qty"] > 0) & (df_all["stock_qty"] <= LOW_STOCK_THRESHOLD)])
    out_s   = len(df_all[df_all["stock_qty"] == 0])
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("รุ่นทั้งหมด",   total_m)
    m2.metric("มีสต๊อกปกติ",   in_s)
    m3.metric("⚠️ ใกล้หมด",   low_s)
    m4.metric("❌ สต๊อกหมด",   out_s)
    st.info("แก้ไขจำนวนสต๊อกในตาราง แล้วกด **บันทึกสต๊อก**")
    edit_cols = ["section","model","btu","price_install","stock_qty","w_install","w_parts","w_comp"]
    edited = st.data_editor(
        df_all[edit_cols].copy(), use_container_width=True, hide_index=True, num_rows="fixed",
        column_config={
            "stock_qty":     st.column_config.NumberColumn("สต๊อก",    min_value=0, step=1),
            "price_install": st.column_config.NumberColumn("ราคา (฿)", min_value=0, step=100),
            "btu":           st.column_config.NumberColumn("BTU",       min_value=0, step=100),
        },
    )
    if st.button("✅ บันทึกสต๊อก", use_container_width=True, type="primary"):
        try: save_stock(clean_df(edited)); st.success("บันทึกสต๊อกแล้ว ✅")
        except Exception as e: st.error(f"ไม่สำเร็จ: {e}")
    if os.path.exists(STOCK_CSV):
        with open(STOCK_CSV,"rb") as f:
            st.download_button("⬇️ Export สต๊อก CSV", data=f.read(),
                               file_name="boonsuk_stock.csv", mime="text/csv")

# ══════════════════════════════════════════════
# PAGE 4: DASHBOARD
# ══════════════════════════════════════════════
elif page == "📊 Dashboard":
    st.title("📊 Dashboard ยอดขาย")
    df_log = load_log()
    if df_log.empty: st.info("ยังไม่มีข้อมูล"); st.stop()

    if "net_total"   in df_log.columns: df_log["net_total"]   = pd.to_numeric(df_log["net_total"],   errors="coerce").fillna(0)
    if "paid_amount" in df_log.columns: df_log["paid_amount"] = pd.to_numeric(df_log["paid_amount"], errors="coerce").fillna(0)

    total_jobs  = len(df_log)
    total_sales = int(df_log["net_total"].sum())
    total_paid  = int(df_log["paid_amount"].sum()) if "paid_amount" in df_log.columns else 0
    avg_sale    = int(df_log["net_total"].mean()) if total_jobs else 0
    jobs_paid   = len(df_log[df_log["status"] == "💰 รับเงินแล้ว"]) if "status" in df_log.columns else 0
    jobs_wait   = len(df_log[df_log["status"] == "📋 รอดำเนินการ"]) if "status" in df_log.columns else 0

    m1,m2,m3,m4,m5,m6 = st.columns(6)
    m1.markdown(f'<div class="metric-card"><h4>งานทั้งหมด</h4><h2>{total_jobs}</h2></div>', unsafe_allow_html=True)
    m2.markdown(f'<div class="metric-card"><h4>ยอดรวม (฿)</h4><h2>{fmt_baht(total_sales)}</h2></div>', unsafe_allow_html=True)
    m3.markdown(f'<div class="metric-card"><h4>รับเงินแล้ว (฿)</h4><h2>{fmt_baht(total_paid)}</h2></div>', unsafe_allow_html=True)
    m4.markdown(f'<div class="metric-card"><h4>เฉลี่ย/งาน (฿)</h4><h2>{fmt_baht(avg_sale)}</h2></div>', unsafe_allow_html=True)
    m5.markdown(f'<div class="metric-card"><h4>รับเงินแล้ว</h4><h2>{jobs_paid} งาน</h2></div>', unsafe_allow_html=True)
    m6.markdown(f'<div class="metric-card"><h4>รอดำเนินการ</h4><h2>{jobs_wait} งาน</h2></div>', unsafe_allow_html=True)

    st.divider()
    ch1, ch2 = st.columns(2)
    with ch1:
        st.subheader("📈 ยอดขายแยกซีรีส์")
        if "section" in df_log.columns:
            bd = df_log.groupby("section")["net_total"].sum().reset_index()
            bd.columns = ["ซีรีส์","ยอดรวม"]
            st.bar_chart(bd.sort_values("ยอดรวม",ascending=False).head(10).set_index("ซีรีส์"))
    with ch2:
        st.subheader("📊 สถานะงาน")
        if "status" in df_log.columns:
            sd = df_log["status"].value_counts().reset_index()
            sd.columns = ["สถานะ","จำนวน"]
            st.bar_chart(sd.set_index("สถานะ"))

    st.divider()

    # ── Export / Import Section ──────────────────
    st.subheader("📁 จัดการข้อมูล (Export / Import)")

    tab1, tab2, tab3 = st.tabs(["📊 Export Excel", "⬇️ Export CSV", "⬆️ Import CSV"])

    with tab1:
        st.caption("Export รายงานยอดขายทั้งหมดเป็น Excel")
        if st.button("📊 สร้างไฟล์ Excel", use_container_width=True, type="primary"):
            xlsx = export_excel(df_log)
            st.download_button("⬇️ ดาวน์โหลด Excel", data=xlsx,
                               file_name=f"รายงาน_บุญสุข_{date.today().strftime('%Y%m%d')}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)

    with tab2:
        st.caption("Export ข้อมูลดิบเป็น CSV สำหรับสำรองข้อมูลหรือแก้ไขภายนอก")
        c1, c2 = st.columns(2)
        with c1:
            st.markdown("**📋 ประวัติงานทั้งหมด**")
            csv_jobs = df_log.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")
            st.download_button(
                "⬇️ ดาวน์โหลด jobs.csv",
                data=csv_jobs,
                file_name=f"boonsuk_jobs_{date.today().strftime('%Y%m%d')}.csv",
                mime="text/csv",
                use_container_width=True
            )
        with c2:
            st.markdown("**📦 สต๊อกสินค้า**")
            df_stock = load_stock()
            csv_stock = df_stock.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")
            st.download_button(
                "⬇️ ดาวน์โหลด stock.csv",
                data=csv_stock,
                file_name=f"boonsuk_stock_{date.today().strftime('%Y%m%d')}.csv",
                mime="text/csv",
                use_container_width=True
            )

    with tab3:
        st.caption("นำเข้าข้อมูลจากไฟล์ CSV (ข้อมูลเดิมจะถูกแทนที่)")
        if st.session_state.get("username") != "admin":
            st.warning("⚠️ เฉพาะ admin เท่านั้นที่สามารถ Import ข้อมูลได้")
        else:
            imp1, imp2 = st.columns(2)
            with imp1:
                st.markdown("**⬆️ Import ประวัติงาน**")
                uploaded_jobs = st.file_uploader("เลือกไฟล์ jobs.csv", type=["csv"], key="upload_jobs")
                if uploaded_jobs:
                    try:
                        df_import = pd.read_csv(uploaded_jobs, encoding="utf-8-sig")
                        st.dataframe(df_import.head(5), use_container_width=True)
                        st.caption(f"พบ {len(df_import)} แถว")
                        if st.button("✅ ยืนยัน Import ประวัติงาน", type="primary", use_container_width=True):
                            if _use_supabase():
                                try:
                                    sb = _get_supabase()
                                    # clear existing then insert
                                    sb.table("jobs").delete().neq("id", 0).execute()
                                    insert_cols = [
                                        "date","customer_name","customer_phone","customer_address",
                                        "section","model","model_btu","base_price","discount",
                                        "extra_install","net_total","paid_amount","status",
                                        "receipt_no","saved_by","room_w","room_l","room_h",
                                        "sun","people","btu","suggest_cap","w_install","w_parts","w_comp"
                                    ]
                                    for _, r in df_import.iterrows():
                                        row = {}
                                        for c in insert_cols:
                                            if c in r:
                                                val = r[c]
                                                if pd.isna(val): val = "" if isinstance(val, str) else 0
                                                row[c] = val
                                        sb.table("jobs").insert(row).execute()
                                    st.success(f"✅ Import สำเร็จ {len(df_import)} รายการ!")
                                    st.rerun()
                                except Exception as e:
                                    st.error(f"Import ไม่สำเร็จ: {e}")
                            else:
                                df_import.to_csv(LOG_CSV, index=False, encoding="utf-8-sig")
                                st.success(f"✅ Import สำเร็จ {len(df_import)} รายการ!")
                                st.rerun()
                    except Exception as e:
                        st.error(f"อ่านไฟล์ไม่สำเร็จ: {e}")

            with imp2:
                st.markdown("**⬆️ Import สต๊อก**")
                uploaded_stock = st.file_uploader("เลือกไฟล์ stock.csv", type=["csv"], key="upload_stock")
                if uploaded_stock:
                    try:
                        df_s_import = pd.read_csv(uploaded_stock, encoding="utf-8-sig")
                        st.dataframe(df_s_import.head(5), use_container_width=True)
                        st.caption(f"พบ {len(df_s_import)} แถว")
                        if st.button("✅ ยืนยัน Import สต๊อก", type="primary", use_container_width=True):
                            save_stock(clean_df(df_s_import))
                            st.success(f"✅ Import สต๊อกสำเร็จ {len(df_s_import)} รายการ!")
                            st.rerun()
                    except Exception as e:
                        st.error(f"อ่านไฟล์ไม่สำเร็จ: {e}")

# ══════════════════════════════════════════════
# PAGE: รับงานซ่อม/บริการ
# ══════════════════════════════════════════════
SERVICE_TYPES = [
    "🆕 ติดตั้งแอร์ใหม่",
    "🔧 ซ่อมแอร์",
    "🚿 ล้างแอร์",
    "🚚 ย้ายแอร์",
    "📡 จานดาวเทียม",
    "❄️ ซ่อมตู้เย็น",
    "👕 ซ่อมเครื่องซักผ้า",
    "📹 ติดตั้ง/ซ่อมกล้องวงจรปิด (CCTV)",
    "📺 ซ่อมทีวี LED/LCD",
]
SERVICE_STATUSES = ["📋 รอดำเนินการ", "🔧 กำลังดำเนินการ", "✅ เสร็จแล้ว", "💰 รับเงินแล้ว", "❌ ยกเลิก"]

def load_service() -> pd.DataFrame:
    if _use_supabase():
        try:
            sb = _get_supabase()
            rows = sb.table("service_jobs").select("*").order("created_at", desc=True).execute().data
            if rows:
                return pd.DataFrame(rows)
            return pd.DataFrame()
        except Exception as e:
            st.warning(f"Supabase load_service: {e}")
    if os.path.exists(SERVICE_CSV):
        try:
            return pd.read_csv(SERVICE_CSV, encoding="utf-8-sig")
        except Exception:
            pass
    return pd.DataFrame()

def save_service(df: pd.DataFrame):
    if _use_supabase():
        try:
            sb = _get_supabase()
            for _, r in df.iterrows():
                row = r.to_dict()
                sb_id = row.get("id")
                if sb_id:
                    sb.table("service_jobs").update({
                        "status": str(row.get("status","")),
                        "note":   str(row.get("note","")),
                        "price":  int(row.get("price",0)),
                    }).eq("id", int(sb_id)).execute()
            return
        except Exception as e:
            st.warning(f"Supabase save_service: {e}")
    df.to_csv(SERVICE_CSV, index=False, encoding="utf-8-sig")

def log_service_job(rec: dict):
    if _use_supabase():
        try:
            sb = _get_supabase()
            insert_cols = ["date","service_type","customer_name","customer_phone",
                           "customer_address","symptom","note","price","status","saved_by"]
            row = {c: rec.get(c,"") for c in insert_cols}
            row["price"] = int(row.get("price",0))
            sb.table("service_jobs").insert(row).execute()
            return
        except Exception as e:
            st.warning(f"Supabase log_service: {e}")
    pd.DataFrame([rec]).to_csv(
        SERVICE_CSV, mode="a",
        header=not os.path.exists(SERVICE_CSV),
        index=False, encoding="utf-8-sig"
    )

if page == "🛠️ รับงานซ่อม/บริการ":
    st.title("🛠️ รับงานซ่อม/บริการ")

    tab_new, tab_list = st.tabs(["➕ เปิดใบงานใหม่", "📋 รายการงานบริการ"])

    # ── Tab 1: เปิดใบงานใหม่ ──────────────────────
    with tab_new:
        st.subheader("📝 กรอกข้อมูลลูกค้า")

        c1, c2 = st.columns(2)
        sv_date   = c1.date_input("📅 วันที่รับงาน", value=date.today())
        sv_type   = c2.selectbox("🔧 ประเภทงาน", SERVICE_TYPES)

        st.divider()
        c3, c4 = st.columns(2)
        sv_name   = c3.text_input("👤 ชื่อลูกค้า", placeholder="ชื่อ-นามสกุล")
        sv_phone  = c4.text_input("📞 เบอร์โทร", placeholder="08X-XXXXXXX")
        sv_addr   = st.text_area("📍 ที่อยู่", placeholder="บ้านเลขที่ หมู่ ตำบล อำเภอ จังหวัด", height=80)

        st.divider()
        sv_symptom = st.text_area("⚡ อาการเสีย / รายละเอียดงาน",
                                   placeholder="เช่น แอร์ไม่เย็น / ขึ้น error E1 / ต้องการล้างแอร์ 2 ตัว",
                                   height=100)
        sv_note   = st.text_area("📌 หมายเหตุเพิ่มเติม (ถ้ามี)", height=60)

        c5, c6 = st.columns(2)
        sv_price  = c5.number_input("💰 ค่าบริการ (บาท)", min_value=0, step=50, value=0)
        sv_status = c6.selectbox("📊 สถานะ", SERVICE_STATUSES)

        st.divider()
        if st.button("💾 บันทึกใบงาน", use_container_width=True, type="primary"):
            if not sv_name.strip():
                st.error("⚠️ กรุณากรอกชื่อลูกค้า")
            elif not sv_phone.strip():
                st.error("⚠️ กรุณากรอกเบอร์โทร")
            elif not sv_symptom.strip():
                st.error("⚠️ กรุณากรอกอาการเสีย/รายละเอียดงาน")
            else:
                rec = {
                    "date":            sv_date.strftime("%Y-%m-%d"),
                    "service_type":    sv_type,
                    "customer_name":   sv_name.strip(),
                    "customer_phone":  sv_phone.strip(),
                    "customer_address":sv_addr.strip(),
                    "symptom":         sv_symptom.strip(),
                    "note":            sv_note.strip(),
                    "price":           int(sv_price),
                    "status":          sv_status,
                    "saved_by":        st.session_state.get("username",""),
                }
                log_service_job(rec)
                st.success(f"✅ บันทึกใบงานสำเร็จ! ลูกค้า: {sv_name.strip()}")
                st.balloons()

                # ── LINE share ──────────────────────────
                price_text = f"{fmt_baht(int(sv_price))} บาท" if int(sv_price) > 0 else "ยังไม่ระบุ"
                line_text = "\n".join([
                    f"🛠️ ใบรับงาน — {STORE_NAME}",
                    f"{'─'*30}",
                    f"📅 วันที่: {sv_date.strftime('%d/%m/%Y')}",
                    f"🔧 ประเภทงาน: {sv_type}",
                    f"{'─'*30}",
                    f"👤 ชื่อ: {sv_name.strip()}",
                    f"📞 เบอร์: {sv_phone.strip()}",
                    f"📍 ที่อยู่: {sv_addr.strip() or '-'}",
                    f"{'─'*30}",
                    f"⚡ อาการ/งาน: {sv_symptom.strip()}",
                    f"📌 หมายเหตุ: {sv_note.strip() or '-'}",
                    f"💰 ค่าบริการ: {price_text}",
                    f"📊 สถานะ: {sv_status}",
                    f"{'─'*30}",
                    f"📞 ติดต่อร้าน: {STORE_PHONE}",
                    f"🌐 {STORE_WEB}",
                ])
                line_url = line_share_link(line_text)

                st.divider()
                st.markdown("### 📤 ส่งใบรับงานให้ลูกค้า")
                bc1, bc2 = st.columns(2)
                with bc1:
                    st.markdown(
                        f'<a href="{line_url}" target="_blank">'
                        f'<button style="width:100%;height:3rem;background:#06C755;color:white;'
                        f'border:none;border-radius:10px;font-size:16px;font-weight:700;cursor:pointer;">'
                        f'💚 ส่ง LINE</button></a>',
                        unsafe_allow_html=True
                    )
                with bc2:
                    st.markdown("**ข้อความที่จะส่ง:**")
                    with st.expander("ดูข้อความ LINE"):
                        st.code(line_text, language=None)

    # ── Tab 2: รายการงาน ─────────────────────────
    with tab_list:
        df_sv = load_service()
        if df_sv.empty:
            st.info("ยังไม่มีใบงานบริการครับ")
        else:
            # Filter bar
            fc1, fc2, fc3 = st.columns([2,2,2])
            sv_filter_status = fc1.selectbox("กรองสถานะ", ["ทั้งหมด"] + SERVICE_STATUSES, key="sv_fs")
            sv_filter_type   = fc2.selectbox("กรองประเภท", ["ทั้งหมด"] + SERVICE_TYPES, key="sv_ft")
            sv_search        = fc3.text_input("🔍 ค้นหาชื่อ/เบอร์", key="sv_srch")

            df_show = df_sv.copy()
            if sv_filter_status != "ทั้งหมด":
                df_show = df_show[df_show["status"] == sv_filter_status]
            if sv_filter_type != "ทั้งหมด":
                df_show = df_show[df_show["service_type"] == sv_filter_type]
            if sv_search:
                mask = (df_show["customer_name"].astype(str).str.contains(sv_search, case=False, na=False) |
                        df_show["customer_phone"].astype(str).str.contains(sv_search, case=False, na=False))
                df_show = df_show[mask]

            st.markdown(f"**พบ {len(df_show)} รายการ**")

            for idx, r in df_show.iterrows():
                status_color = {
                    "📋 รอดำเนินการ": "#1565c0",
                    "🔧 กำลังดำเนินการ": "#e65100",
                    "✅ เสร็จแล้ว": "#2e7d32",
                    "💰 รับเงินแล้ว": "#6a1b9a",
                    "❌ ยกเลิก": "#c62828",
                }.get(str(r.get("status","")), "#888")

                with st.expander(
                    f"**{r.get('service_type','')}** — {r.get('customer_name','')} | {r.get('customer_phone','')} | {r.get('date','')}",
                    expanded=False
                ):
                    d1, d2 = st.columns(2)
                    d1.markdown(f"**📍 ที่อยู่:** {r.get('customer_address','-')}")
                    d2.markdown(f"**💰 ค่าบริการ:** {fmt_baht(r.get('price',0))} บาท")
                    st.markdown(f"**⚡ อาการ:** {r.get('symptom','-')}")
                    if r.get("note",""):
                        st.markdown(f"**📌 หมายเหตุ:** {r.get('note','')}")
                    st.markdown(f"<span style='color:{status_color};font-weight:700'>● {r.get('status','')}</span>",
                                unsafe_allow_html=True)

                    # อัปเดตสถานะ
                    u1, u2, u3 = st.columns([3,2,1])
                    new_st = u1.selectbox("เปลี่ยนสถานะ", SERVICE_STATUSES,
                                          index=SERVICE_STATUSES.index(str(r.get("status", SERVICE_STATUSES[0])))
                                          if str(r.get("status","")) in SERVICE_STATUSES else 0,
                                          key=f"sv_st_{idx}")
                    new_price = u2.number_input("ค่าบริการ", min_value=0, step=50,
                                                value=int(r.get("price",0)), key=f"sv_pr_{idx}")
                    if u3.button("💾", key=f"sv_save_{idx}", help="บันทึก"):
                        df_sv.at[idx, "status"] = new_st
                        df_sv.at[idx, "price"]  = new_price
                        save_service(df_sv)
                        st.success("✅ อัปเดตแล้ว"); st.rerun()

            st.divider()
            # Export CSV
            csv_sv = df_show.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")
            st.download_button("⬇️ Export CSV ใบงานบริการ", data=csv_sv,
                               file_name=f"service_jobs_{date.today().strftime('%Y%m%d')}.csv",
                               mime="text/csv", use_container_width=True)

# ══════════════════════════════════════════════
# ERROR CODE DATABASE
# ══════════════════════════════════════════════
ERROR_DB = {
    "Daikin": [
        {"code":"A1","desc":"PCB แผงวงจรผิดปกติ","cause":"PCB เสีย / ไฟเลี้ยงกระชาก","fix":"ตรวจสอบแรงดันไฟ 220V, เปลี่ยน PCB ชุดในห้อง","parts":"PCB indoor"},
        {"code":"A3","desc":"ระดับน้ำทิ้งผิดปกติ / float switch","cause":"ท่อน้ำทิ้งอุดตัน / float switch เสีย","fix":"ล้างท่อน้ำทิ้ง, ตรวจ float switch","parts":"Float switch"},
        {"code":"A5","desc":"เทอร์มิสเตอร์ freeze protection ทำงาน","cause":"แผง evaporator สกปรก / น้ำยาน้อย","fix":"ล้างแผงอีวา, ตรวจน้ำยา","parts":"น้ำยาแอร์ R32/R410A"},
        {"code":"A6","desc":"มอเตอร์พัดลม indoor ผิดปกติ","cause":"มอเตอร์เสีย / capacitor เสีย","fix":"ตรวจ capacitor, เปลี่ยนมอเตอร์","parts":"มอเตอร์พัดลม indoor, Capacitor"},
        {"code":"A7","desc":"แผ่น swing motor ผิดปกติ","cause":"มอเตอร์ swing เสีย","fix":"เปลี่ยนมอเตอร์ swing","parts":"Step motor swing"},
        {"code":"C4","desc":"เทอร์มิสเตอร์ liquid pipe ผิดปกติ","cause":"เทอร์มิสเตอร์ขาด/ช็อต","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor liquid"},
        {"code":"C5","desc":"เทอร์มิสเตอร์ gas pipe ผิดปกติ","cause":"เทอร์มิสเตอร์ขาด/ช็อต","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor gas"},
        {"code":"C9","desc":"เทอร์มิสเตอร์อุณหภูมิห้องผิดปกติ","cause":"เทอร์มิสเตอร์เสีย","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor room temp"},
        {"code":"E1","desc":"PCB outdoor ผิดปกติ","cause":"PCB outdoor เสีย","fix":"ตรวจไฟ outdoor, เปลี่ยน PCB","parts":"PCB outdoor"},
        {"code":"E3","desc":"แรงดันสูง high pressure","cause":"คอยล์ร้อนสกปรก / พัดลมเสีย / น้ำยาเกิน","fix":"ล้างคอยล์ร้อน, ตรวจพัดลม, ตรวจน้ำยา","parts":"พัดลม outdoor, High pressure switch"},
        {"code":"E4","desc":"แรงดันต่ำ low pressure","cause":"น้ำยารั่ว/น้อย / ท่ออุดตัน","fix":"ตรวจน้ำยา, ตรวจท่อ","parts":"น้ำยาแอร์"},
        {"code":"E5","desc":"กระแสไฟ compressor เกิน (overcurrent)","cause":"แรงดันไฟต่ำ / compressor เสีย","fix":"ตรวจแรงดันไฟ, ตรวจ compressor","parts":"Compressor"},
        {"code":"E6","desc":"การสื่อสาร indoor-outdoor ผิดปกติ","cause":"สายสัญญาณขาด/หลวม","fix":"ตรวจสายสัญญาณ 3 เส้น","parts":"สายสัญญาณ"},
        {"code":"E7","desc":"มอเตอร์พัดลม outdoor ผิดปกติ","cause":"มอเตอร์เสีย / inverter board เสีย","fix":"ตรวจมอเตอร์, ตรวจ IPM","parts":"มอเตอร์ outdoor, IPM module"},
        {"code":"E9","desc":"วาล์ว 4 ทาง (4-way valve) ผิดปกติ","cause":"4-way valve เสีย / คอยล์วาล์วเสีย","fix":"ตรวจคอยล์วาล์ว, เปลี่ยนวาล์ว","parts":"4-way valve"},
        {"code":"F3","desc":"เทอร์มิสเตอร์ discharge pipe ผิดปกติ","cause":"เทอร์มิสเตอร์เสีย","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor discharge"},
        {"code":"H6","desc":"ตรวจไม่พบตำแหน่ง rotor","cause":"IPM เสีย / compressor เสีย","fix":"ตรวจ IPM module, ตรวจ compressor","parts":"IPM module, Compressor"},
        {"code":"H9","desc":"เทอร์มิสเตอร์อุณหภูมิภายนอกผิดปกติ","cause":"เทอร์มิสเตอร์เสีย","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor outdoor temp"},
        {"code":"J3","desc":"เทอร์มิสเตอร์ discharge temperature สูง","cause":"น้ำยาน้อย / ท่ออุดตัน","fix":"ตรวจน้ำยา, ตรวจท่อ","parts":"น้ำยาแอร์"},
        {"code":"L4","desc":"IPM module อุณหภูมิสูง","cause":"ระบายความร้อนไม่ดี","fix":"ล้างคอยล์ร้อน, ตรวจพัดลม outdoor","parts":"IPM module"},
        {"code":"L5","desc":"IPM module overcurrent","cause":"IPM เสีย / compressor ล็อก","fix":"เปลี่ยน IPM, ตรวจ compressor","parts":"IPM module, Compressor"},
        {"code":"U0","desc":"น้ำยาแอร์น้อย","cause":"น้ำยารั่ว","fix":"หาจุดรั่ว, เติมน้ำยา","parts":"น้ำยาแอร์"},
        {"code":"U2","desc":"แรงดันไฟ DC ผิดปกติ","cause":"ไฟบ้านผิดปกติ / PCB เสีย","fix":"ตรวจแรงดันไฟ, เปลี่ยน PCB","parts":"PCB outdoor"},
        {"code":"U4","desc":"การสื่อสาร indoor-outdoor ผิดปกติ","cause":"สายสัญญาณขาด","fix":"ตรวจสายสัญญาณ","parts":"สายสัญญาณ"},
    ],
    "Mitsubishi Electric": [
        {"code":"P1","desc":"แรงดันไฟ indoor ผิดปกติ","cause":"ไฟตก/เกิน หรือ PCB เสีย","fix":"ตรวจแรงดันไฟ 220V, เปลี่ยน PCB","parts":"PCB indoor"},
        {"code":"P2","desc":"แรงดันไฟ outdoor ผิดปกติ","cause":"ไฟตก/เกิน","fix":"ตรวจแรงดันไฟ, ตรวจ capacitor","parts":"Capacitor, PCB outdoor"},
        {"code":"P4","desc":"เซ็นเซอร์ drain ผิดปกติ / น้ำล้น","cause":"ท่อน้ำทิ้งอุดตัน","fix":"ล้างท่อน้ำทิ้ง, ตรวจ float sensor","parts":"Float sensor"},
        {"code":"E0","desc":"อุปกรณ์ป้องกัน (protection device) ทำงาน","cause":"แรงดันสูง/ต่ำ, อุณหภูมิสูง","fix":"ตรวจน้ำยา, ล้างคอยล์","parts":"High/low pressure switch"},
        {"code":"E1","desc":"PCB outdoor ผิดปกติ","cause":"PCB outdoor เสีย","fix":"เปลี่ยน PCB outdoor","parts":"PCB outdoor"},
        {"code":"E2","desc":"Zero-cross signal ผิดปกติ","cause":"PCB เสีย","fix":"เปลี่ยน PCB indoor","parts":"PCB indoor"},
        {"code":"E3","desc":"Fan motor ผิดปกติ (indoor)","cause":"มอเตอร์เสีย / ใบพัดติดขัด","fix":"ตรวจใบพัด, เปลี่ยนมอเตอร์","parts":"มอเตอร์พัดลม indoor"},
        {"code":"E4","desc":"Fan motor ผิดปกติ (outdoor)","cause":"มอเตอร์เสีย","fix":"เปลี่ยนมอเตอร์ outdoor","parts":"มอเตอร์พัดลม outdoor"},
        {"code":"E5","desc":"กระแสไฟ compressor เกิน","cause":"compressor ล็อก / ไฟต่ำ","fix":"ตรวจแรงดันไฟ, เปลี่ยน compressor","parts":"Compressor"},
        {"code":"E6","desc":"Compressor หยุดผิดปกติ","cause":"น้ำยาน้อย / compressor เสีย","fix":"ตรวจน้ำยา, ตรวจ compressor","parts":"Compressor"},
        {"code":"E7","desc":"Fan motor outdoor หยุดผิดปกติ","cause":"มอเตอร์เสีย","fix":"เปลี่ยนมอเตอร์","parts":"มอเตอร์พัดลม outdoor"},
        {"code":"E8","desc":"กระแสไฟ input เกิน","cause":"ไฟกระชาก","fix":"ตรวจแรงดันไฟ, ตรวจ PCB","parts":"PCB outdoor"},
        {"code":"E9","desc":"4-way valve ผิดปกติ","cause":"4-way valve เสีย","fix":"เปลี่ยน 4-way valve","parts":"4-way valve"},
        {"code":"F1","desc":"เซ็นเซอร์อุณหภูมิห้อง (indoor) เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor room"},
        {"code":"F2","desc":"เซ็นเซอร์ pipe temp (indoor) เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor pipe indoor"},
        {"code":"F3","desc":"เซ็นเซอร์ outdoor temp เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor outdoor"},
        {"code":"F4","desc":"เซ็นเซอร์ discharge temp เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor discharge"},
        {"code":"F5","desc":"เซ็นเซอร์ outdoor pipe temp เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor outdoor pipe"},
        {"code":"P5","desc":"อุณหภูมิ inverter สูงเกิน","cause":"ระบายความร้อน IPM ไม่ดี","fix":"ล้างคอยล์ร้อน, ตรวจพัดลม","parts":"IPM module"},
        {"code":"P6","desc":"ตรวจพบความผิดปกติ inverter","cause":"IPM เสีย","fix":"เปลี่ยน IPM module","parts":"IPM module"},
        {"code":"P8","desc":"อุณหภูมิ outdoor สูงเกิน","cause":"คอยล์ร้อนสกปรก","fix":"ล้างคอยล์ร้อน","parts":"-"},
    ],
    "Mitsubishi Heavy": [
        {"code":"E0","desc":"Protection device ทำงาน","cause":"High/low pressure trip","fix":"ตรวจน้ำยา, ล้างคอยล์","parts":"Pressure switch"},
        {"code":"E1","desc":"PCB indoor ผิดปกติ","cause":"PCB เสีย","fix":"เปลี่ยน PCB indoor","parts":"PCB indoor"},
        {"code":"E3","desc":"มอเตอร์พัดลม indoor ผิดปกติ","cause":"มอเตอร์/capacitor เสีย","fix":"ตรวจ capacitor, เปลี่ยนมอเตอร์","parts":"มอเตอร์ indoor, Capacitor"},
        {"code":"E5","desc":"Compressor overcurrent","cause":"ไฟต่ำ / compressor เสีย","fix":"ตรวจไฟ, เปลี่ยน compressor","parts":"Compressor"},
        {"code":"E6","desc":"Compressor หยุดผิดปกติ","cause":"น้ำยาน้อย / compressor เสีย","fix":"ตรวจน้ำยา","parts":"Compressor"},
        {"code":"E9","desc":"4-way valve ผิดปกติ","cause":"4-way valve เสีย","fix":"เปลี่ยน 4-way valve","parts":"4-way valve"},
        {"code":"F1","desc":"เซ็นเซอร์อุณหภูมิห้องเสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor room"},
        {"code":"F2","desc":"เซ็นเซอร์ pipe indoor เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor pipe"},
        {"code":"F3","desc":"เซ็นเซอร์ outdoor temp เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor outdoor"},
        {"code":"F4","desc":"เซ็นเซอร์ discharge เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor discharge"},
        {"code":"P1","desc":"High pressure trip","cause":"คอยล์ร้อนสกปรก / น้ำยาเกิน","fix":"ล้างคอยล์ร้อน, ตรวจน้ำยา","parts":"High pressure switch"},
        {"code":"P2","desc":"Low pressure trip","cause":"น้ำยารั่ว/น้อย","fix":"ตรวจน้ำยา, หาจุดรั่ว","parts":"Low pressure switch, น้ำยา"},
        {"code":"P4","desc":"ท่อน้ำทิ้งเต็ม / drain error","cause":"ท่อน้ำทิ้งอุดตัน","fix":"ล้างท่อน้ำทิ้ง","parts":"Float switch"},
        {"code":"P5","desc":"IPM อุณหภูมิสูง","cause":"IPM ระบายร้อนไม่ดี","fix":"ล้างคอยล์, ตรวจพัดลม outdoor","parts":"IPM module"},
        {"code":"P6","desc":"Inverter ผิดปกติ","cause":"IPM เสีย","fix":"เปลี่ยน IPM","parts":"IPM module"},
        {"code":"P8","desc":"อุณหภูมิ outdoor สูง","cause":"คอยล์ร้อนสกปรก","fix":"ล้างคอยล์ร้อน","parts":"-"},
        {"code":"P9","desc":"4-way valve ผิดปกติ","cause":"คอยล์วาล์วเสีย","fix":"ตรวจ/เปลี่ยน 4-way valve","parts":"4-way valve"},
    ],
    "Carrier": [
        {"code":"E1","desc":"Indoor PCB ผิดปกติ","cause":"PCB เสีย / ไฟผิดปกติ","fix":"ตรวจไฟ, เปลี่ยน PCB indoor","parts":"PCB indoor"},
        {"code":"E2","desc":"Outdoor PCB ผิดปกติ","cause":"PCB outdoor เสีย","fix":"เปลี่ยน PCB outdoor","parts":"PCB outdoor"},
        {"code":"E3","desc":"High pressure protection","cause":"คอยล์ร้อนสกปรก / น้ำยาเกิน","fix":"ล้างคอยล์ร้อน, ตรวจน้ำยา","parts":"High pressure switch"},
        {"code":"E4","desc":"Low pressure protection","cause":"น้ำยาน้อย/รั่ว","fix":"ตรวจน้ำยา, หาจุดรั่ว","parts":"Low pressure switch, น้ำยา"},
        {"code":"E5","desc":"Compressor overcurrent","cause":"ไฟต่ำ / compressor เสีย","fix":"ตรวจไฟ, ตรวจ compressor","parts":"Compressor"},
        {"code":"E6","desc":"การสื่อสาร indoor-outdoor ผิดปกติ","cause":"สายสัญญาณขาด","fix":"ตรวจสายสัญญาณ","parts":"สายสัญญาณ"},
        {"code":"E7","desc":"มอเตอร์พัดลม indoor ผิดปกติ","cause":"มอเตอร์/capacitor เสีย","fix":"ตรวจ capacitor, เปลี่ยนมอเตอร์","parts":"มอเตอร์ indoor, Capacitor"},
        {"code":"E8","desc":"มอเตอร์พัดลม outdoor ผิดปกติ","cause":"มอเตอร์เสีย","fix":"เปลี่ยนมอเตอร์ outdoor","parts":"มอเตอร์ outdoor"},
        {"code":"E9","desc":"Freeze protection (คอยล์เย็นแข็ง)","cause":"แผงสกปรก / น้ำยาน้อย","fix":"ล้างแผง, ตรวจน้ำยา","parts":"น้ำยาแอร์"},
        {"code":"F1","desc":"เซ็นเซอร์อุณหภูมิห้องเสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor room"},
        {"code":"F2","desc":"เซ็นเซอร์ indoor pipe เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor pipe"},
        {"code":"F3","desc":"เซ็นเซอร์ discharge เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor discharge"},
        {"code":"F4","desc":"เซ็นเซอร์ outdoor temp เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor outdoor"},
        {"code":"H6","desc":"Drain pump / float switch ผิดปกติ","cause":"ท่อน้ำทิ้งอุดตัน","fix":"ล้างท่อ, ตรวจ float switch","parts":"Float switch"},
        {"code":"P1","desc":"IPM / inverter protection","cause":"IPM เสีย","fix":"เปลี่ยน IPM module","parts":"IPM module"},
        {"code":"P2","desc":"Voltage ผิดปกติ","cause":"ไฟตก/เกิน","fix":"ตรวจแรงดันไฟ","parts":"PCB outdoor"},
    ],
    "Fujitsu": [
        {"code":"00","desc":"ปกติ (ไม่มี error)","cause":"-","fix":"-","parts":"-"},
        {"code":"03","desc":"Fan motor indoor ผิดปกติ","cause":"มอเตอร์/capacitor เสีย","fix":"ตรวจ capacitor, เปลี่ยนมอเตอร์","parts":"มอเตอร์ indoor, Capacitor"},
        {"code":"04","desc":"Drain pump / float switch ผิดปกติ","cause":"ท่อน้ำทิ้งอุดตัน","fix":"ล้างท่อน้ำทิ้ง","parts":"Float switch"},
        {"code":"05","desc":"Freeze protection","cause":"แผงสกปรก / น้ำยาน้อย","fix":"ล้างแผง, ตรวจน้ำยา","parts":"น้ำยาแอร์"},
        {"code":"06","desc":"High pressure trip","cause":"คอยล์ร้อนสกปรก","fix":"ล้างคอยล์ร้อน","parts":"High pressure switch"},
        {"code":"07","desc":"Low pressure trip","cause":"น้ำยาน้อย","fix":"ตรวจน้ำยา","parts":"น้ำยาแอร์"},
        {"code":"08","desc":"Compressor overcurrent","cause":"compressor เสีย / ไฟต่ำ","fix":"ตรวจไฟ, ตรวจ compressor","parts":"Compressor"},
        {"code":"09","desc":"Inverter / IPM ผิดปกติ","cause":"IPM เสีย","fix":"เปลี่ยน IPM module","parts":"IPM module"},
        {"code":"12","desc":"การสื่อสาร indoor-outdoor ผิดปกติ","cause":"สายสัญญาณขาด","fix":"ตรวจสายสัญญาณ","parts":"สายสัญญาณ"},
        {"code":"13","desc":"PCB outdoor ผิดปกติ","cause":"PCB เสีย","fix":"เปลี่ยน PCB outdoor","parts":"PCB outdoor"},
        {"code":"14","desc":"อุณหภูมิ discharge สูงเกิน","cause":"น้ำยาน้อย / ท่ออุดตัน","fix":"ตรวจน้ำยา","parts":"น้ำยาแอร์"},
        {"code":"15","desc":"เซ็นเซอร์อุณหภูมิห้องเสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor room"},
        {"code":"16","desc":"เซ็นเซอร์ pipe indoor เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor pipe"},
        {"code":"17","desc":"เซ็นเซอร์ outdoor temp เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor outdoor"},
        {"code":"18","desc":"เซ็นเซอร์ discharge temp เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor discharge"},
        {"code":"19","desc":"Fan motor outdoor ผิดปกติ","cause":"มอเตอร์เสีย","fix":"เปลี่ยนมอเตอร์ outdoor","parts":"มอเตอร์ outdoor"},
        {"code":"20","desc":"4-way valve ผิดปกติ","cause":"4-way valve เสีย","fix":"เปลี่ยน 4-way valve","parts":"4-way valve"},
        {"code":"22","desc":"IPM อุณหภูมิสูง","cause":"ระบายความร้อนไม่ดี","fix":"ล้างคอยล์, ตรวจพัดลม","parts":"IPM module"},
    ],
    "Gree": [
        {"code":"C5","desc":"เซ็นเซอร์อุณหภูมิห้องเสีย","cause":"เทอร์มิสเตอร์ขาด/ช็อต","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor room"},
        {"code":"C4","desc":"เซ็นเซอร์ pipe indoor เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor pipe indoor"},
        {"code":"C1","desc":"เซ็นเซอร์ outdoor temp เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor outdoor"},
        {"code":"C3","desc":"เซ็นเซอร์ discharge เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor discharge"},
        {"code":"E1","desc":"High pressure protection","cause":"คอยล์ร้อนสกปรก","fix":"ล้างคอยล์ร้อน","parts":"High pressure switch"},
        {"code":"E2","desc":"High temperature discharge protection","cause":"น้ำยาน้อย","fix":"ตรวจน้ำยา","parts":"น้ำยาแอร์"},
        {"code":"E3","desc":"Low pressure protection","cause":"น้ำยารั่ว","fix":"หาจุดรั่ว, เติมน้ำยา","parts":"น้ำยาแอร์"},
        {"code":"E4","desc":"Compressor discharge temp สูงเกิน","cause":"น้ำยาน้อย","fix":"ตรวจน้ำยา","parts":"น้ำยาแอร์"},
        {"code":"E5","desc":"Overcurrent protection","cause":"ไฟต่ำ / compressor เสีย","fix":"ตรวจไฟ, ตรวจ compressor","parts":"Compressor"},
        {"code":"E6","desc":"การสื่อสาร indoor-outdoor ผิดปกติ","cause":"สายสัญญาณขาด","fix":"ตรวจสายสัญญาณ","parts":"สายสัญญาณ"},
        {"code":"F0","desc":"IPM ผิดปกติ","cause":"IPM เสีย","fix":"เปลี่ยน IPM module","parts":"IPM module"},
        {"code":"F1","desc":"Compressor phase loss","cause":"ไฟ 3 เฟสผิดปกติ (เฉพาะรุ่นใหญ่)","fix":"ตรวจไฟ 3 เฟส","parts":"-"},
        {"code":"H3","desc":"Compressor overload protection","cause":"compressor ร้อนเกิน","fix":"ตรวจน้ำยา, ตรวจการระบายความร้อน","parts":"Overload protector"},
        {"code":"H6","desc":"ตรวจพบตำแหน่ง motor ไม่ได้","cause":"IPM เสีย / compressor เสีย","fix":"ตรวจ IPM, ตรวจ compressor","parts":"IPM module, Compressor"},
        {"code":"Hc","desc":"มอเตอร์พัดลม outdoor ผิดปกติ","cause":"มอเตอร์เสีย","fix":"เปลี่ยนมอเตอร์ outdoor","parts":"มอเตอร์ outdoor"},
        {"code":"Ld","desc":"มอเตอร์พัดลม indoor ผิดปกติ","cause":"มอเตอร์/capacitor เสีย","fix":"ตรวจ capacitor, เปลี่ยนมอเตอร์","parts":"มอเตอร์ indoor, Capacitor"},
        {"code":"LP","desc":"Low pressure switch trip","cause":"น้ำยาน้อย","fix":"ตรวจน้ำยา","parts":"Low pressure switch"},
    ],
    "Midea": [
        {"code":"E1","desc":"High pressure protection","cause":"คอยล์ร้อนสกปรก / น้ำยาเกิน","fix":"ล้างคอยล์ร้อน, ตรวจน้ำยา","parts":"High pressure switch"},
        {"code":"E2","desc":"Freeze protection indoor coil","cause":"แผงสกปรก / น้ำยาน้อย","fix":"ล้างแผง, ตรวจน้ำยา","parts":"น้ำยาแอร์"},
        {"code":"E3","desc":"Low pressure protection","cause":"น้ำยาน้อย/รั่ว","fix":"หาจุดรั่ว, เติมน้ำยา","parts":"น้ำยาแอร์"},
        {"code":"E4","desc":"High discharge temp protection","cause":"น้ำยาน้อย / ท่ออุดตัน","fix":"ตรวจน้ำยา","parts":"น้ำยาแอร์"},
        {"code":"E5","desc":"Overcurrent protection","cause":"ไฟต่ำ / compressor เสีย","fix":"ตรวจไฟ, ตรวจ compressor","parts":"Compressor"},
        {"code":"E6","desc":"การสื่อสาร indoor-outdoor ผิดปกติ","cause":"สายสัญญาณขาด","fix":"ตรวจสายสัญญาณ","parts":"สายสัญญาณ"},
        {"code":"E7","desc":"PCB indoor ผิดปกติ","cause":"PCB เสีย","fix":"เปลี่ยน PCB indoor","parts":"PCB indoor"},
        {"code":"E8","desc":"กระแสไฟ input เกิน","cause":"ไฟกระชาก","fix":"ตรวจแรงดันไฟ","parts":"PCB outdoor"},
        {"code":"E9","desc":"4-way valve ผิดปกติ","cause":"4-way valve เสีย","fix":"เปลี่ยน 4-way valve","parts":"4-way valve"},
        {"code":"F0","desc":"IPM ผิดปกติ","cause":"IPM เสีย","fix":"เปลี่ยน IPM module","parts":"IPM module"},
        {"code":"F1","desc":"เซ็นเซอร์อุณหภูมิห้องเสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor room"},
        {"code":"F2","desc":"เซ็นเซอร์ pipe indoor เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor pipe"},
        {"code":"F3","desc":"เซ็นเซอร์ outdoor temp เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor outdoor"},
        {"code":"F4","desc":"เซ็นเซอร์ discharge เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor discharge"},
        {"code":"P0","desc":"IPM protection","cause":"IPM เสีย","fix":"เปลี่ยน IPM","parts":"IPM module"},
        {"code":"P1","desc":"Voltage protection","cause":"ไฟตก/เกิน","fix":"ตรวจแรงดันไฟ","parts":"PCB outdoor"},
        {"code":"P2","desc":"เซ็นเซอร์ outdoor pipe เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor outdoor pipe"},
    ],
    "MAVELL": [
        {"code":"E1","desc":"High pressure protection","cause":"คอยล์ร้อนสกปรก","fix":"ล้างคอยล์ร้อน","parts":"High pressure switch"},
        {"code":"E2","desc":"Freeze / low temp protection","cause":"แผงสกปรก / น้ำยาน้อย","fix":"ล้างแผง, ตรวจน้ำยา","parts":"น้ำยาแอร์"},
        {"code":"E3","desc":"Low pressure protection","cause":"น้ำยารั่ว","fix":"หาจุดรั่ว, เติมน้ำยา","parts":"น้ำยาแอร์"},
        {"code":"E4","desc":"Compressor discharge temp สูง","cause":"น้ำยาน้อย","fix":"ตรวจน้ำยา","parts":"น้ำยาแอร์"},
        {"code":"E5","desc":"Overcurrent protection","cause":"compressor เสีย / ไฟต่ำ","fix":"ตรวจไฟ, ตรวจ compressor","parts":"Compressor"},
        {"code":"E6","desc":"การสื่อสาร indoor-outdoor ผิดปกติ","cause":"สายสัญญาณขาด","fix":"ตรวจสายสัญญาณ","parts":"สายสัญญาณ"},
        {"code":"F1","desc":"เซ็นเซอร์อุณหภูมิห้องเสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor room"},
        {"code":"F2","desc":"เซ็นเซอร์ pipe เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor pipe"},
        {"code":"P1","desc":"IPM protection","cause":"IPM เสีย","fix":"เปลี่ยน IPM","parts":"IPM module"},
        {"code":"P2","desc":"Voltage protection","cause":"ไฟตก/เกิน","fix":"ตรวจแรงดันไฟ","parts":"PCB outdoor"},
    ],
    "Samsung": [
        {"code":"C4 / C5","desc":"เซ็นเซอร์อุณหภูมิเสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor"},
        {"code":"E1","desc":"High pressure protection","cause":"คอยล์ร้อนสกปรก","fix":"ล้างคอยล์ร้อน","parts":"High pressure switch"},
        {"code":"E2","desc":"Freeze protection","cause":"แผงสกปรก / น้ำยาน้อย","fix":"ล้างแผง, ตรวจน้ำยา","parts":"น้ำยาแอร์"},
        {"code":"E3","desc":"Low pressure protection","cause":"น้ำยาน้อย","fix":"ตรวจน้ำยา","parts":"น้ำยาแอร์"},
        {"code":"E4","desc":"Discharge temp สูง","cause":"น้ำยาน้อย","fix":"ตรวจน้ำยา","parts":"น้ำยาแอร์"},
        {"code":"E5","desc":"Overcurrent protection","cause":"compressor เสีย","fix":"ตรวจ compressor","parts":"Compressor"},
        {"code":"E6","desc":"การสื่อสาร ผิดปกติ","cause":"สายสัญญาณขาด","fix":"ตรวจสายสัญญาณ","parts":"สายสัญญาณ"},
        {"code":"E7","desc":"PCB ผิดปกติ","cause":"PCB เสีย","fix":"เปลี่ยน PCB","parts":"PCB"},
        {"code":"E8","desc":"มอเตอร์พัดลม indoor ผิดปกติ","cause":"มอเตอร์เสีย","fix":"เปลี่ยนมอเตอร์","parts":"มอเตอร์ indoor"},
        {"code":"E9","desc":"4-way valve ผิดปกติ","cause":"4-way valve เสีย","fix":"เปลี่ยน 4-way valve","parts":"4-way valve"},
        {"code":"F1","desc":"PCB outdoor เสีย","cause":"PCB เสีย","fix":"เปลี่ยน PCB outdoor","parts":"PCB outdoor"},
        {"code":"Cl","desc":"ต้องล้างแผง (clean filter)","cause":"แผงกรองสกปรก","fix":"ล้างแผงกรอง","parts":"-"},
    ],
    "LG": [
        {"code":"C1","desc":"เซ็นเซอร์อุณหภูมิห้องเสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor room"},
        {"code":"C2","desc":"เซ็นเซอร์ pipe indoor เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor pipe"},
        {"code":"C3","desc":"เซ็นเซอร์ outdoor temp เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor outdoor"},
        {"code":"C5","desc":"เซ็นเซอร์ discharge เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor discharge"},
        {"code":"CH01","desc":"เซ็นเซอร์อุณหภูมิห้องผิดปกติ","cause":"เทอร์มิสเตอร์เสีย","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor room"},
        {"code":"CH02","desc":"เซ็นเซอร์ pipe indoor ผิดปกติ","cause":"เทอร์มิสเตอร์เสีย","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor pipe"},
        {"code":"CH03","desc":"เซ็นเซอร์ outdoor ผิดปกติ","cause":"เทอร์มิสเตอร์เสีย","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor outdoor"},
        {"code":"CH05","desc":"การสื่อสาร indoor-outdoor ผิดปกติ","cause":"สายสัญญาณขาด","fix":"ตรวจสายสัญญาณ","parts":"สายสัญญาณ"},
        {"code":"CH06","desc":"Fan motor indoor ผิดปกติ","cause":"มอเตอร์/capacitor เสีย","fix":"ตรวจ capacitor, เปลี่ยนมอเตอร์","parts":"มอเตอร์ indoor, Capacitor"},
        {"code":"CH07","desc":"Fan motor outdoor ผิดปกติ","cause":"มอเตอร์เสีย","fix":"เปลี่ยนมอเตอร์ outdoor","parts":"มอเตอร์ outdoor"},
        {"code":"CH09","desc":"Inverter ผิดปกติ","cause":"IPM เสีย","fix":"เปลี่ยน IPM module","parts":"IPM module"},
        {"code":"CH10","desc":"DC peak current protection","cause":"compressor เสีย","fix":"ตรวจ compressor","parts":"Compressor"},
        {"code":"CH21","desc":"High pressure protection","cause":"คอยล์ร้อนสกปรก","fix":"ล้างคอยล์ร้อน","parts":"High pressure switch"},
        {"code":"CH22","desc":"Low pressure protection","cause":"น้ำยาน้อย","fix":"ตรวจน้ำยา","parts":"น้ำยาแอร์"},
        {"code":"CH23","desc":"Discharge temp สูงเกิน","cause":"น้ำยาน้อย","fix":"ตรวจน้ำยา","parts":"น้ำยาแอร์"},
        {"code":"CH24","desc":"Compressor overcurrent","cause":"compressor เสีย / ไฟต่ำ","fix":"ตรวจไฟ, ตรวจ compressor","parts":"Compressor"},
        {"code":"CH25","desc":"4-way valve ผิดปกติ","cause":"4-way valve เสีย","fix":"เปลี่ยน 4-way valve","parts":"4-way valve"},
        {"code":"CH38","desc":"Drain pump / float switch ผิดปกติ","cause":"ท่อน้ำทิ้งอุดตัน","fix":"ล้างท่อน้ำทิ้ง","parts":"Float switch"},
    ],
    "Beko": [
        {"code":"E0","desc":"PCB indoor ผิดปกติ","cause":"PCB เสีย / ไฟผิดปกติ","fix":"ตรวจไฟ 220V, เปลี่ยน PCB indoor","parts":"PCB indoor"},
        {"code":"E1","desc":"เซ็นเซอร์อุณหภูมิห้องเสีย","cause":"เทอร์มิสเตอร์ขาด/ช็อต","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor room"},
        {"code":"E2","desc":"เซ็นเซอร์ pipe indoor เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor pipe indoor"},
        {"code":"E3","desc":"เซ็นเซอร์ outdoor temp เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor outdoor"},
        {"code":"E4","desc":"เซ็นเซอร์ discharge temp เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor discharge"},
        {"code":"E5","desc":"High pressure protection","cause":"คอยล์ร้อนสกปรก / น้ำยาเกิน","fix":"ล้างคอยล์ร้อน, ตรวจน้ำยา","parts":"High pressure switch"},
        {"code":"E6","desc":"Low pressure protection","cause":"น้ำยาน้อย/รั่ว","fix":"หาจุดรั่ว, เติมน้ำยา","parts":"Low pressure switch, น้ำยาแอร์"},
        {"code":"E7","desc":"Compressor overcurrent","cause":"compressor เสีย / ไฟต่ำ","fix":"ตรวจแรงดันไฟ, ตรวจ compressor","parts":"Compressor"},
        {"code":"E8","desc":"การสื่อสาร indoor-outdoor ผิดปกติ","cause":"สายสัญญาณขาด/หลวม","fix":"ตรวจสายสัญญาณ","parts":"สายสัญญาณ"},
        {"code":"E9","desc":"มอเตอร์พัดลม indoor ผิดปกติ","cause":"มอเตอร์/capacitor เสีย","fix":"ตรวจ capacitor, เปลี่ยนมอเตอร์","parts":"มอเตอร์ indoor, Capacitor"},
        {"code":"F1","desc":"Freeze protection (คอยล์เย็นแข็ง)","cause":"แผงสกปรก / น้ำยาน้อย","fix":"ล้างแผง evaporator, ตรวจน้ำยา","parts":"น้ำยาแอร์"},
        {"code":"F2","desc":"Drain / float switch ผิดปกติ","cause":"ท่อน้ำทิ้งอุดตัน","fix":"ล้างท่อน้ำทิ้ง, ตรวจ float switch","parts":"Float switch"},
        {"code":"P1","desc":"IPM / inverter protection","cause":"IPM เสีย","fix":"เปลี่ยน IPM module","parts":"IPM module"},
        {"code":"P2","desc":"Voltage protection","cause":"ไฟตก/เกิน","fix":"ตรวจแรงดันไฟ","parts":"PCB outdoor"},
    ],
    "Haier": [
        {"code":"E1","desc":"เซ็นเซอร์อุณหภูมิห้องเสีย","cause":"เทอร์มิสเตอร์ขาด/ช็อต","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor room"},
        {"code":"E2","desc":"เซ็นเซอร์ indoor pipe เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor pipe indoor"},
        {"code":"E3","desc":"เซ็นเซอร์ outdoor temp เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor outdoor"},
        {"code":"E4","desc":"เซ็นเซอร์ discharge เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor discharge"},
        {"code":"E5","desc":"Freeze protection","cause":"แผงสกปรก / น้ำยาน้อย","fix":"ล้างแผง, ตรวจน้ำยา","parts":"น้ำยาแอร์"},
        {"code":"E6","desc":"การสื่อสาร indoor-outdoor ผิดปกติ","cause":"สายสัญญาณขาด","fix":"ตรวจสายสัญญาณ","parts":"สายสัญญาณ"},
        {"code":"E7","desc":"EEPROM PCB ผิดปกติ","cause":"PCB เสีย","fix":"เปลี่ยน PCB indoor","parts":"PCB indoor"},
        {"code":"E9","desc":"มอเตอร์ swing ผิดปกติ","cause":"step motor เสีย","fix":"เปลี่ยน step motor swing","parts":"Step motor swing"},
        {"code":"F1","desc":"High pressure protection","cause":"คอยล์ร้อนสกปรก","fix":"ล้างคอยล์ร้อน","parts":"High pressure switch"},
        {"code":"F2","desc":"Low pressure protection","cause":"น้ำยาน้อย/รั่ว","fix":"หาจุดรั่ว, เติมน้ำยา","parts":"น้ำยาแอร์"},
        {"code":"F3","desc":"Compressor overcurrent","cause":"compressor เสีย / ไฟต่ำ","fix":"ตรวจไฟ, ตรวจ compressor","parts":"Compressor"},
        {"code":"F4","desc":"Discharge temp สูงเกิน","cause":"น้ำยาน้อย","fix":"ตรวจน้ำยา","parts":"น้ำยาแอร์"},
        {"code":"F5","desc":"มอเตอร์พัดลม indoor ผิดปกติ","cause":"มอเตอร์/capacitor เสีย","fix":"ตรวจ capacitor, เปลี่ยนมอเตอร์","parts":"มอเตอร์ indoor, Capacitor"},
        {"code":"F6","desc":"มอเตอร์พัดลม outdoor ผิดปกติ","cause":"มอเตอร์เสีย","fix":"เปลี่ยนมอเตอร์ outdoor","parts":"มอเตอร์ outdoor"},
        {"code":"P1","desc":"Voltage ผิดปกติ","cause":"ไฟตก/เกิน","fix":"ตรวจแรงดันไฟ","parts":"PCB outdoor"},
        {"code":"P2","desc":"IPM protection","cause":"IPM เสีย","fix":"เปลี่ยน IPM module","parts":"IPM module"},
        {"code":"P3","desc":"4-way valve ผิดปกติ","cause":"4-way valve เสีย","fix":"เปลี่ยน 4-way valve","parts":"4-way valve"},
    ],
    "Toshiba": [
        {"code":"E01","desc":"เซ็นเซอร์อุณหภูมิห้อง (indoor) เสีย","cause":"เทอร์มิสเตอร์ขาด/ช็อต","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor room"},
        {"code":"E02","desc":"เซ็นเซอร์ pipe indoor เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor pipe indoor"},
        {"code":"E03","desc":"เซ็นเซอร์ outdoor ambient เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor outdoor"},
        {"code":"E04","desc":"เซ็นเซอร์ discharge เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor discharge"},
        {"code":"E05","desc":"เซ็นเซอร์ outdoor pipe เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor outdoor pipe"},
        {"code":"E06","desc":"PCB indoor ผิดปกติ","cause":"PCB เสีย","fix":"เปลี่ยน PCB indoor","parts":"PCB indoor"},
        {"code":"E07","desc":"PCB outdoor ผิดปกติ","cause":"PCB เสีย","fix":"เปลี่ยน PCB outdoor","parts":"PCB outdoor"},
        {"code":"E08","desc":"การสื่อสาร indoor-outdoor ผิดปกติ","cause":"สายสัญญาณขาด","fix":"ตรวจสายสัญญาณ","parts":"สายสัญญาณ"},
        {"code":"E09","desc":"High pressure protection","cause":"คอยล์ร้อนสกปรก / น้ำยาเกิน","fix":"ล้างคอยล์ร้อน, ตรวจน้ำยา","parts":"High pressure switch"},
        {"code":"E10","desc":"Low pressure protection","cause":"น้ำยาน้อย/รั่ว","fix":"หาจุดรั่ว, เติมน้ำยา","parts":"น้ำยาแอร์"},
        {"code":"E11","desc":"Compressor overcurrent","cause":"compressor เสีย / ไฟต่ำ","fix":"ตรวจไฟ, ตรวจ compressor","parts":"Compressor"},
        {"code":"E12","desc":"Discharge temp สูงเกิน","cause":"น้ำยาน้อย","fix":"ตรวจน้ำยา","parts":"น้ำยาแอร์"},
        {"code":"E13","desc":"มอเตอร์พัดลม indoor ผิดปกติ","cause":"มอเตอร์/capacitor เสีย","fix":"ตรวจ capacitor, เปลี่ยนมอเตอร์","parts":"มอเตอร์ indoor, Capacitor"},
        {"code":"E14","desc":"มอเตอร์พัดลม outdoor ผิดปกติ","cause":"มอเตอร์เสีย","fix":"เปลี่ยนมอเตอร์ outdoor","parts":"มอเตอร์ outdoor"},
        {"code":"E15","desc":"Inverter / IPM protection","cause":"IPM เสีย","fix":"เปลี่ยน IPM module","parts":"IPM module"},
        {"code":"E16","desc":"4-way valve ผิดปกติ","cause":"4-way valve เสีย","fix":"เปลี่ยน 4-way valve","parts":"4-way valve"},
        {"code":"E17","desc":"Freeze protection","cause":"แผงสกปรก / น้ำยาน้อย","fix":"ล้างแผง, ตรวจน้ำยา","parts":"น้ำยาแอร์"},
        {"code":"E18","desc":"Drain pump / float switch ผิดปกติ","cause":"ท่อน้ำทิ้งอุดตัน","fix":"ล้างท่อน้ำทิ้ง","parts":"Float switch"},
    ],
    "Panasonic": [
        {"code":"H00","desc":"ปกติ ไม่มีปัญหา","cause":"-","fix":"-","parts":"-"},
        {"code":"H11","desc":"การสื่อสาร indoor-outdoor ผิดปกติ","cause":"สายสัญญาณขาด/หลวม","fix":"ตรวจสายสัญญาณ 3 เส้น","parts":"สายสัญญาณ"},
        {"code":"H12","desc":"ความจุ outdoor ไม่ตรงกับ indoor","cause":"ติดตั้งผิดรุ่น","fix":"ตรวจสอบรุ่น indoor-outdoor ให้ตรงกัน","parts":"-"},
        {"code":"H14","desc":"เซ็นเซอร์ outdoor temp เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor outdoor"},
        {"code":"H15","desc":"เซ็นเซอร์ compressor temp เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor compressor"},
        {"code":"H16","desc":"กระแสไฟ outdoor ต่ำผิดปกติ","cause":"น้ำยาน้อย / compressor เสียบางส่วน","fix":"ตรวจน้ำยา, ตรวจ compressor","parts":"Compressor"},
        {"code":"H19","desc":"มอเตอร์พัดลม indoor ผิดปกติ","cause":"มอเตอร์/capacitor เสีย","fix":"ตรวจ capacitor, เปลี่ยนมอเตอร์","parts":"มอเตอร์ indoor, Capacitor"},
        {"code":"H23","desc":"เซ็นเซอร์ pipe indoor เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor pipe indoor"},
        {"code":"H27","desc":"High pressure switch ทำงาน","cause":"คอยล์ร้อนสกปรก / น้ำยาเกิน","fix":"ล้างคอยล์ร้อน, ตรวจน้ำยา","parts":"High pressure switch"},
        {"code":"H28","desc":"เซ็นเซอร์ outdoor pipe เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor outdoor pipe"},
        {"code":"H33","desc":"Phase sequence ผิดปกติ (3 เฟส)","cause":"ไฟ 3 เฟสสลับเฟส","fix":"สลับสายไฟ L1/L2","parts":"-"},
        {"code":"H36","desc":"เซ็นเซอร์ discharge เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor discharge"},
        {"code":"H38","desc":"เซ็นเซอร์อุณหภูมิห้อง (indoor) เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor room"},
        {"code":"F90","desc":"Freeze protection (high pressure)","cause":"น้ำยาเกิน / คอยล์ร้อนสกปรก","fix":"ล้างคอยล์ร้อน, ตรวจน้ำยา","parts":"High pressure switch"},
        {"code":"F91","desc":"Freeze protection (low pressure)","cause":"น้ำยาน้อย/รั่ว","fix":"หาจุดรั่ว, เติมน้ำยา","parts":"น้ำยาแอร์"},
        {"code":"F93","desc":"Compressor overcurrent / overload","cause":"compressor เสีย","fix":"ตรวจ compressor","parts":"Compressor, Overload protector"},
        {"code":"F95","desc":"High pressure protection","cause":"คอยล์ร้อนสกปรก","fix":"ล้างคอยล์ร้อน","parts":"High pressure switch"},
        {"code":"F96","desc":"IPM / inverter overheat","cause":"IPM ระบายร้อนไม่ดี","fix":"ล้างคอยล์, ตรวจพัดลม","parts":"IPM module"},
        {"code":"F97","desc":"Compressor discharge temp สูงเกิน","cause":"น้ำยาน้อย","fix":"ตรวจน้ำยา","parts":"น้ำยาแอร์"},
        {"code":"F98","desc":"Total input power สูงเกิน","cause":"ไฟกระชาก","fix":"ตรวจแรงดันไฟ","parts":"PCB outdoor"},
        {"code":"F99","desc":"DC current protection","cause":"IPM เสีย / compressor เสีย","fix":"เปลี่ยน IPM, ตรวจ compressor","parts":"IPM module, Compressor"},
    ],
    "AUX": [
        {"code":"E1","desc":"High pressure protection","cause":"คอยล์ร้อนสกปรก / น้ำยาเกิน","fix":"ล้างคอยล์ร้อน, ตรวจน้ำยา","parts":"High pressure switch"},
        {"code":"E2","desc":"Freeze protection indoor coil","cause":"แผงสกปรก / น้ำยาน้อย","fix":"ล้างแผง, ตรวจน้ำยา","parts":"น้ำยาแอร์"},
        {"code":"E3","desc":"Low pressure protection","cause":"น้ำยาน้อย/รั่ว","fix":"หาจุดรั่ว, เติมน้ำยา","parts":"น้ำยาแอร์"},
        {"code":"E4","desc":"Discharge temp สูงเกิน","cause":"น้ำยาน้อย / ท่ออุดตัน","fix":"ตรวจน้ำยา, ตรวจท่อ","parts":"น้ำยาแอร์"},
        {"code":"E5","desc":"Overcurrent protection","cause":"compressor เสีย / ไฟต่ำ","fix":"ตรวจไฟ, ตรวจ compressor","parts":"Compressor"},
        {"code":"E6","desc":"การสื่อสาร indoor-outdoor ผิดปกติ","cause":"สายสัญญาณขาด","fix":"ตรวจสายสัญญาณ","parts":"สายสัญญาณ"},
        {"code":"E7","desc":"มอเตอร์พัดลม indoor ผิดปกติ","cause":"มอเตอร์/capacitor เสีย","fix":"ตรวจ capacitor, เปลี่ยนมอเตอร์","parts":"มอเตอร์ indoor, Capacitor"},
        {"code":"E8","desc":"มอเตอร์พัดลม outdoor ผิดปกติ","cause":"มอเตอร์เสีย","fix":"เปลี่ยนมอเตอร์ outdoor","parts":"มอเตอร์ outdoor"},
        {"code":"E9","desc":"4-way valve ผิดปกติ","cause":"4-way valve เสีย","fix":"เปลี่ยน 4-way valve","parts":"4-way valve"},
        {"code":"F1","desc":"เซ็นเซอร์อุณหภูมิห้องเสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor room"},
        {"code":"F2","desc":"เซ็นเซอร์ pipe indoor เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor pipe"},
        {"code":"F3","desc":"เซ็นเซอร์ outdoor temp เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor outdoor"},
        {"code":"F4","desc":"เซ็นเซอร์ discharge เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor discharge"},
        {"code":"P0","desc":"IPM protection","cause":"IPM เสีย","fix":"เปลี่ยน IPM module","parts":"IPM module"},
        {"code":"P1","desc":"Voltage ผิดปกติ","cause":"ไฟตก/เกิน","fix":"ตรวจแรงดันไฟ","parts":"PCB outdoor"},
        {"code":"H3","desc":"Compressor overload protection","cause":"compressor ร้อนเกิน","fix":"ตรวจน้ำยา, ตรวจระบายความร้อน","parts":"Overload protector"},
        {"code":"H6","desc":"ตรวจพบตำแหน่ง motor ไม่ได้","cause":"IPM เสีย / compressor เสีย","fix":"ตรวจ IPM, ตรวจ compressor","parts":"IPM module, Compressor"},
    ],
    "Hisense": [
        {"code":"E1","desc":"High pressure protection","cause":"คอยล์ร้อนสกปรก / พัดลมเสีย","fix":"ล้างคอยล์ร้อน, ตรวจพัดลม","parts":"High pressure switch"},
        {"code":"E2","desc":"Freeze protection indoor coil","cause":"แผงสกปรก / น้ำยาน้อย","fix":"ล้างแผง, ตรวจน้ำยา","parts":"น้ำยาแอร์"},
        {"code":"E3","desc":"Low pressure protection","cause":"น้ำยาน้อย/รั่ว","fix":"หาจุดรั่ว, เติมน้ำยา","parts":"น้ำยาแอร์"},
        {"code":"E4","desc":"Discharge temp สูงเกิน","cause":"น้ำยาน้อย","fix":"ตรวจน้ำยา","parts":"น้ำยาแอร์"},
        {"code":"E5","desc":"Overcurrent protection","cause":"compressor เสีย / ไฟต่ำ","fix":"ตรวจไฟ, ตรวจ compressor","parts":"Compressor"},
        {"code":"E6","desc":"การสื่อสาร indoor-outdoor ผิดปกติ","cause":"สายสัญญาณขาด","fix":"ตรวจสายสัญญาณ","parts":"สายสัญญาณ"},
        {"code":"E7","desc":"PCB indoor ผิดปกติ","cause":"PCB เสีย","fix":"เปลี่ยน PCB indoor","parts":"PCB indoor"},
        {"code":"E8","desc":"มอเตอร์พัดลม indoor ผิดปกติ","cause":"มอเตอร์/capacitor เสีย","fix":"ตรวจ capacitor, เปลี่ยนมอเตอร์","parts":"มอเตอร์ indoor, Capacitor"},
        {"code":"E9","desc":"มอเตอร์พัดลม outdoor ผิดปกติ","cause":"มอเตอร์เสีย","fix":"เปลี่ยนมอเตอร์ outdoor","parts":"มอเตอร์ outdoor"},
        {"code":"F1","desc":"เซ็นเซอร์อุณหภูมิห้องเสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor room"},
        {"code":"F2","desc":"เซ็นเซอร์ pipe indoor เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor pipe"},
        {"code":"F3","desc":"เซ็นเซอร์ outdoor เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor outdoor"},
        {"code":"F4","desc":"เซ็นเซอร์ discharge เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor discharge"},
        {"code":"P0","desc":"IPM / inverter protection","cause":"IPM เสีย","fix":"เปลี่ยน IPM module","parts":"IPM module"},
        {"code":"P1","desc":"Voltage ผิดปกติ","cause":"ไฟตก/เกิน","fix":"ตรวจแรงดันไฟ","parts":"PCB outdoor"},
        {"code":"H6","desc":"ตรวจพบตำแหน่ง rotor ไม่ได้","cause":"IPM เสีย / compressor เสีย","fix":"ตรวจ IPM, ตรวจ compressor","parts":"IPM module, Compressor"},
        {"code":"Ld","desc":"Drain / float switch ผิดปกติ","cause":"ท่อน้ำทิ้งอุดตัน","fix":"ล้างท่อน้ำทิ้ง","parts":"Float switch"},
    ],
    "Saijo Denki": [
        {"code":"E1","desc":"เซ็นเซอร์อุณหภูมิห้อง (indoor) เสีย","cause":"เทอร์มิสเตอร์ขาด/ช็อต","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor room"},
        {"code":"E2","desc":"เซ็นเซอร์ pipe indoor เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor pipe indoor"},
        {"code":"E3","desc":"เซ็นเซอร์ outdoor temp เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor outdoor"},
        {"code":"E4","desc":"เซ็นเซอร์ discharge เสีย","cause":"เทอร์มิสเตอร์ขาด","fix":"เปลี่ยนเทอร์มิสเตอร์","parts":"Thermistor discharge"},
        {"code":"E5","desc":"High pressure protection","cause":"คอยล์ร้อนสกปรก / น้ำยาเกิน","fix":"ล้างคอยล์ร้อน, ตรวจน้ำยา","parts":"High pressure switch"},
        {"code":"E6","desc":"Low pressure / น้ำยาน้อย","cause":"น้ำยารั่ว/น้อย","fix":"หาจุดรั่ว, เติมน้ำยา","parts":"น้ำยาแอร์"},
        {"code":"E7","desc":"Compressor overcurrent","cause":"compressor เสีย / ไฟต่ำ","fix":"ตรวจไฟ, ตรวจ compressor","parts":"Compressor"},
        {"code":"E8","desc":"การสื่อสาร indoor-outdoor ผิดปกติ","cause":"สายสัญญาณขาด/หลวม","fix":"ตรวจสายสัญญาณ","parts":"สายสัญญาณ"},
        {"code":"E9","desc":"PCB indoor ผิดปกติ","cause":"PCB เสีย / ไฟผิดปกติ","fix":"ตรวจไฟ, เปลี่ยน PCB indoor","parts":"PCB indoor"},
        {"code":"F1","desc":"มอเตอร์พัดลม indoor ผิดปกติ","cause":"มอเตอร์/capacitor เสีย","fix":"ตรวจ capacitor, เปลี่ยนมอเตอร์","parts":"มอเตอร์ indoor, Capacitor"},
        {"code":"F2","desc":"มอเตอร์พัดลม outdoor ผิดปกติ","cause":"มอเตอร์เสีย","fix":"เปลี่ยนมอเตอร์ outdoor","parts":"มอเตอร์ outdoor"},
        {"code":"F3","desc":"Freeze protection","cause":"แผงสกปรก / น้ำยาน้อย","fix":"ล้างแผง evaporator, ตรวจน้ำยา","parts":"น้ำยาแอร์"},
        {"code":"F4","desc":"Drain / float switch ผิดปกติ","cause":"ท่อน้ำทิ้งอุดตัน","fix":"ล้างท่อน้ำทิ้ง, ตรวจ float switch","parts":"Float switch"},
        {"code":"F5","desc":"4-way valve ผิดปกติ","cause":"4-way valve เสีย","fix":"เปลี่ยน 4-way valve","parts":"4-way valve"},
        {"code":"P1","desc":"IPM / inverter protection","cause":"IPM เสีย","fix":"เปลี่ยน IPM module","parts":"IPM module"},
        {"code":"P2","desc":"Voltage protection","cause":"ไฟตก/เกิน","fix":"ตรวจแรงดันไฟ","parts":"PCB outdoor"},
    ],
}

# ══════════════════════════════════════════════
# PAGE 5: ERROR CODE LIBRARY
# ══════════════════════════════════════════════
if page == "🔧 คลังเออเร่อแอร์":
    st.title("🔧 คลังเออเร่อโค้ดแอร์")
    st.caption("รวม error code แอร์ทุกยี่ห้อ พร้อมสาเหตุ วิธีแก้ไข และอะไหล่ที่ต้องเปลี่ยน")

    # ── search bar ──────────────────────────────
    srch = st.text_input("🔍 พิมพ์ error code หรือคำค้นหา", placeholder="เช่น E1, compressor, น้ำยา, discharge").strip().lower()

    # ── brand filter ────────────────────────────
    brands = list(ERROR_DB.keys())
    sel_brand = st.selectbox("เลือกยี่ห้อ", ["ทั้งหมด"] + brands)

    # ── build filtered table ────────────────────
    rows = []
    for brand, errors in ERROR_DB.items():
        if sel_brand != "ทั้งหมด" and brand != sel_brand:
            continue
        for e in errors:
            rows.append({
                "ยี่ห้อ": brand,
                "Code": e["code"],
                "ความหมาย": e["desc"],
                "สาเหตุ": e["cause"],
                "วิธีแก้ไข": e["fix"],
                "อะไหล่ที่ต้องเปลี่ยน": e["parts"],
            })
    df_err = pd.DataFrame(rows)

    if srch:
        mask = pd.Series(False, index=df_err.index)
        for col in df_err.columns:
            mask |= df_err[col].astype(str).str.lower().str.contains(srch, na=False)
        df_err = df_err[mask]

    st.markdown(f"**พบ {len(df_err)} รายการ**")

    if df_err.empty:
        st.warning("ไม่พบ error code ที่ค้นหา")
    else:
        # แสดงแบบ card ถ้าเลือกยี่ห้อเดียวหรือค้นหา
        if sel_brand != "ทั้งหมด" or srch:
            for _, r in df_err.iterrows():
                parts_color = "#c62828" if r["อะไหล่ที่ต้องเปลี่ยน"] not in ["-",""] else "#388e3c"
                st.markdown(f"""
                <div style="background:#fff;border-radius:10px;padding:14px 18px;margin-bottom:10px;
                            border-left:5px solid #1565c0;box-shadow:0 1px 4px #0001;">
                  <div style="display:flex;justify-content:space-between;align-items:center;">
                    <span style="font-size:22px;font-weight:900;color:#1565c0;">{r['Code']}</span>
                    <span style="background:#e3f2fd;color:#1565c0;padding:2px 10px;border-radius:20px;font-size:13px;">{r['ยี่ห้อ']}</span>
                  </div>
                  <div style="font-size:15px;font-weight:700;margin:4px 0;">📋 {r['ความหมาย']}</div>
                  <div style="font-size:13px;color:#555;">⚡ <b>สาเหตุ:</b> {r['สาเหตุ']}</div>
                  <div style="font-size:13px;color:#2e7d32;">🔧 <b>วิธีแก้ไข:</b> {r['วิธีแก้ไข']}</div>
                  <div style="font-size:13px;color:{parts_color};">🔩 <b>อะไหล่:</b> {r['อะไหล่ที่ต้องเปลี่ยน']}</div>
                </div>
                """, unsafe_allow_html=True)
        else:
            # แสดงแบบตารางถ้าดูทั้งหมด
            st.dataframe(df_err, use_container_width=True, hide_index=True,
                         column_config={
                             "Code": st.column_config.TextColumn("Code", width="small"),
                             "ยี่ห้อ": st.column_config.TextColumn("ยี่ห้อ", width="small"),
                         })

    # ── export ──────────────────────────────────
    st.divider()
    if st.button("📥 Export คลังเออเร่อ Excel", use_container_width=True):
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:  # requires: pip install openpyxl
            all_rows = []
            for brand, errors in ERROR_DB.items():
                for e in errors:
                    all_rows.append({"ยี่ห้อ":brand,"Code":e["code"],"ความหมาย":e["desc"],
                                     "สาเหตุ":e["cause"],"วิธีแก้ไข":e["fix"],"อะไหล่":e["parts"]})
            pd.DataFrame(all_rows).to_excel(writer, sheet_name="Error Codes", index=False)
            wb = writer.book
            from openpyxl.styles import Font, PatternFill, Alignment
            ws = wb["Error Codes"]
            for cell in ws[1]:
                cell.fill = PatternFill("solid", fgColor="B71C1C")
                cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
                cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                mlen = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(mlen + 4, 50)
        st.download_button("⬇️ ดาวน์โหลด Excel คลังเออเร่อ", data=buf.getvalue(),
                           file_name="คลังเออเร่อแอร์_บุญสุข.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)

# ══════════════════════════════════════════════
# PAGE 6: IMPORT / EXPORT (admin only)
# ══════════════════════════════════════════════
elif page == "⚙️ นำเข้า/ส่งออกข้อมูล":
    if st.session_state.get("username") != "admin":
        st.error("เฉพาะ admin เท่านั้น"); st.stop()

    st.title("⚙️ นำเข้า/ส่งออกข้อมูล")
    st.caption("สำรอง-กู้คืนข้อมูลทั้งหมด | เฉพาะผู้ดูแลระบบ")

    tab1, tab2 = st.tabs(["📤 ส่งออก (Export)", "📥 นำเข้า (Import)"])

    # ──────────────────────────────────────────
    # TAB 1: EXPORT
    # ──────────────────────────────────────────
    with tab1:
        st.subheader("📤 ดาวน์โหลดข้อมูลทั้งหมด")

        col1, col2 = st.columns(2)

        # Export ประวัติงาน
        with col1:
            st.markdown("#### 📋 ประวัติงาน / ใบเสนอราคา")
            df_log = load_log()
            if df_log.empty:
                st.info("ยังไม่มีข้อมูลงาน")
            else:
                st.metric("จำนวนงานทั้งหมด", f"{len(df_log)} รายการ")
                # CSV
                csv_log = df_log.to_csv(index=False, encoding="utf-8-sig")
                st.download_button(
                    "⬇️ Export ประวัติงาน (.csv)",
                    data=csv_log.encode("utf-8-sig"),
                    file_name=f"jobs_backup_{date.today().strftime('%Y%m%d')}.csv",
                    mime="text/csv",
                    use_container_width=True,
                )
                # Excel
                xlsx_bytes = export_excel(df_log)
                st.download_button(
                    "⬇️ Export ประวัติงาน (.xlsx)",
                    data=xlsx_bytes,
                    file_name=f"jobs_backup_{date.today().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

        # Export สต๊อก
        with col2:
            st.markdown("#### 📦 ข้อมูลสต๊อกแอร์")
            df_stk = load_stock()
            st.metric("จำนวนรุ่นทั้งหมด", f"{len(df_stk)} รุ่น")
            csv_stk = df_stk.to_csv(index=False, encoding="utf-8-sig")
            st.download_button(
                "⬇️ Export สต๊อก (.csv)",
                data=csv_stk.encode("utf-8-sig"),
                file_name=f"stock_backup_{date.today().strftime('%Y%m%d')}.csv",
                mime="text/csv",
                use_container_width=True,
            )

        # Export ทุกอย่างใน ZIP
        st.divider()
        st.markdown("#### 📦 Export ทุกอย่างในครั้งเดียว")
        if st.button("⬇️ Download ทั้งหมด (ZIP)", use_container_width=True, type="primary"):
            import zipfile as zf
            zip_buf = io.BytesIO()
            with zf.ZipFile(zip_buf, "w") as z:
                df_log2 = load_log()
                df_stk2 = load_stock()
                if not df_log2.empty:
                    z.writestr(f"jobs_{date.today().strftime('%Y%m%d')}.csv",
                               df_log2.to_csv(index=False, encoding="utf-8-sig"))
                z.writestr(f"stock_{date.today().strftime('%Y%m%d')}.csv",
                           df_stk2.to_csv(index=False, encoding="utf-8-sig"))
            st.download_button(
                "⬇️ ดาวน์โหลด ZIP",
                data=zip_buf.getvalue(),
                file_name=f"boonsuk_backup_{date.today().strftime('%Y%m%d')}.zip",
                mime="application/zip",
                use_container_width=True,
            )

    # ──────────────────────────────────────────
    # TAB 2: IMPORT
    # ──────────────────────────────────────────
    with tab2:
        st.subheader("📥 นำเข้าข้อมูลจากไฟล์ CSV")
        st.warning("⚠️ การนำเข้าจะ **เพิ่ม** ข้อมูลเข้าไปในระบบ ไม่ลบข้อมูลเดิม")

        imp1, imp2 = st.columns(2)

        # Import ประวัติงาน
        with imp1:
            st.markdown("#### 📋 นำเข้าประวัติงาน")
            uploaded_jobs = st.file_uploader(
                "เลือกไฟล์ CSV ประวัติงาน",
                type=["csv"],
                key="upload_jobs",
            )
            if uploaded_jobs:
                try:
                    df_new = pd.read_csv(uploaded_jobs, encoding="utf-8-sig")
                    st.success(f"พบข้อมูล {len(df_new)} รายการ")
                    st.dataframe(df_new.head(5), use_container_width=True, hide_index=True)
                    if st.button("✅ ยืนยันนำเข้าประวัติงาน", use_container_width=True, type="primary", key="confirm_jobs"):
                        if _use_supabase():
                            sb = _get_supabase()
                            insert_cols = [
                                "date","customer_name","customer_phone","customer_address",
                                "section","model","model_btu","base_price","discount",
                                "extra_install","net_total","paid_amount","status",
                                "receipt_no","saved_by","room_w","room_l","room_h",
                                "sun","people","btu","suggest_cap","w_install","w_parts","w_comp"
                            ]
                            ok = 0
                            for _, r in df_new.iterrows():
                                try:
                                    row = {}
                                    for c in insert_cols:
                                        if c in r:
                                            val = r[c]
                                            if pd.isna(val): val = ""
                                            if isinstance(val, float): val = int(val) if val == int(val) else val
                                            row[c] = val
                                    sb.table("jobs").insert(row).execute()
                                    ok += 1
                                except Exception:
                                    pass
                            st.success(f"นำเข้าสำเร็จ {ok}/{len(df_new)} รายการ ✅")
                        else:
                            # CSV mode: append
                            existing = load_log()
                            combined = pd.concat([existing, df_new], ignore_index=True) if not existing.empty else df_new
                            save_log(combined)
                            st.success(f"นำเข้าสำเร็จ {len(df_new)} รายการ ✅")
                        st.cache_data.clear()
                except Exception as e:
                    st.error(f"อ่านไฟล์ไม่ได้: {e}")

        # Import สต๊อก
        with imp2:
            st.markdown("#### 📦 นำเข้า/อัปเดตสต๊อก")
            uploaded_stock = st.file_uploader(
                "เลือกไฟล์ CSV สต๊อก",
                type=["csv"],
                key="upload_stock",
            )
            if uploaded_stock:
                try:
                    df_stk_new = pd.read_csv(uploaded_stock, encoding="utf-8-sig")
                    df_stk_new = clean_df(df_stk_new)
                    st.success(f"พบข้อมูล {len(df_stk_new)} รุ่น")
                    st.dataframe(df_stk_new.head(5), use_container_width=True, hide_index=True)
                    if st.button("✅ ยืนยันอัปเดตสต๊อก", use_container_width=True, type="primary", key="confirm_stock"):
                        save_stock(df_stk_new)
                        st.success("อัปเดตสต๊อกสำเร็จ ✅")
                        st.cache_data.clear()
                except Exception as e:
                    st.error(f"อ่านไฟล์ไม่ได้: {e}")

        # ── ลบข้อมูลทั้งหมด (Danger Zone) ────────
        st.divider()
        with st.expander("🔴 Danger Zone — ลบข้อมูลทั้งหมด"):
            st.error("⚠️ การลบข้อมูลไม่สามารถกู้คืนได้!")
            confirm_del = st.text_input("พิมพ์ 'ลบทั้งหมด' เพื่อยืนยัน")
            if st.button("🗑️ ลบประวัติงานทั้งหมด", use_container_width=True):
                if confirm_del == "ลบทั้งหมด":
                    if _use_supabase():
                        try:
                            sb = _get_supabase()
                            sb.table("jobs").delete().neq("id", 0).execute()
                            st.success("ลบข้อมูลทั้งหมดแล้ว ✅")
                            st.cache_data.clear()
                        except Exception as e:
                            st.error(f"ไม่สำเร็จ: {e}")
                    else:
                        if os.path.exists(LOG_CSV):
                            os.remove(LOG_CSV)
                            st.success("ลบไฟล์ CSV แล้ว ✅")
                else:
                    st.warning("พิมพ์ 'ลบทั้งหมด' ให้ถูกต้องก่อนครับ")
